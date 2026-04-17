# Copyright (c) 2026 BelTech Systems LLC and contributors
# SPDX-License-Identifier: MIT
"""tests/test_results_collector.py — Unit tests for agents/results_collector.py.

Phase: v1a (OI-29)
Layer: Tests

Tests parse_xsim_log(), write_vpr_results(), and generate_gap_report_json()
using synthetic openpyxl workbooks and synthetic log files. The real BALU
VPR is never written to — all VPR operations use tmp_path fixtures only.
"""
import json
import os

import openpyxl
import pytest

from agents.results_collector import (
    SimResult,
    generate_gap_report_json,
    parse_xsim_log,
    write_vpr_results,
)


# ── Synthetic workbook helpers ────────────────────────────────────────────────

# Column layout used by all test workbooks.
# Mirrors the real VPR column positions so parse_vplan() (hardcoded 0-based
# indices) and write_vpr_results() (dynamic header scan) both work.
#
#  0  Req_ID
#  1  Family
#  3  Statement
#  6  Verification_Method
#  7  Covered_By
# 10  Disposition
# 11  Waiver_Rationale
# 18  RTL_Status
# 19  RTL_Run_Date
# 20  RTL_Commit
# 21  RTL_Evidence
# 27  Overall_Status  (formula placeholder — never written by pssgen)
_N_COLS = 28  # minimum column count for the test workbooks


def _make_header_row() -> list:
    """Build a 28-element row of header strings at the real VPR positions."""
    row = [None] * _N_COLS
    row[0]  = "Req_ID"
    row[1]  = "Family"
    row[3]  = "Statement"
    row[6]  = "Verification_Method"
    row[7]  = "Covered_By"
    row[10] = "Disposition"
    row[11] = "Waiver_Rationale"
    row[18] = "RTL_Status"
    row[19] = "RTL_Run_Date"
    row[20] = "RTL_Commit"
    row[21] = "RTL_Evidence"
    row[27] = "Overall_Status"
    return row


def _make_data_row(
    req_id: str,
    disposition: str = "GENERATED",
    rtl_status: str = "NOT_RUN",
    family: str = "FAM",
) -> list:
    """Build a 28-element VPR data row with minimal fields populated."""
    row = [None] * _N_COLS
    row[0]  = req_id
    row[1]  = family
    row[3]  = f"The IP shall {req_id}."
    row[7]  = "COV-001"
    row[10] = disposition
    row[18] = rtl_status
    return row


def _make_vpr_workbook(data_rows: list[list], vpr_path: str) -> str:
    """Write a minimal VPR workbook to vpr_path and return the path.

    The workbook has:
      - VPR tab: group-headers row, column-headers row, example row, data rows
      - Coverage_Goals tab: header row only (no data needed for most tests)

    Args:
        data_rows: List of 28-element row lists to append as VPR data rows.
        vpr_path: Destination .xlsx path (must end with .xlsx).

    Returns:
        vpr_path.
    """
    wb = openpyxl.Workbook()
    ws_vpr = wb.active
    ws_vpr.title = "VPR"

    # Row 1: group headers (content irrelevant — parser skips row 1)
    ws_vpr.append(["Group headers"] + [None] * (_N_COLS - 1))
    # Row 2: column headers — must be present for dynamic column lookup
    ws_vpr.append(_make_header_row())
    # Row 3: example/template placeholder (parser skips row 3)
    ws_vpr.append(["[BLOCK-FAM-NNN]"] + [None] * (_N_COLS - 1))
    # Data rows
    for row in data_rows:
        padded = list(row) + [None] * (_N_COLS - len(row))
        ws_vpr.append(padded)

    # Coverage_Goals tab — required by parse_vplan()
    ws_cov = wb.create_sheet("Coverage_Goals")
    ws_cov.append(
        ["ID", "Name", "Description", "Stimulus_Strategy",
         "Boundary_Values", "Linked_Requirements", "Status",
         "Coverage_Type", "Notes"]
    )
    # Add a minimal COV item so the tab is not empty
    ws_cov.append(
        ["COV-001", "Basic coverage", "Covers the nominal path", "Write then read",
         "", "TEST-001", "Open", "Functional", ""]
    )

    wb.save(vpr_path)
    return vpr_path


def _make_passing_sim_result(log_path: str = "xsim.log") -> SimResult:
    """Return a SimResult representing a clean (PASS) simulation run."""
    return SimResult(
        log_path=log_path,
        passed=True,
        uvm_errors=0,
        uvm_fatals=0,
        uvm_warnings=0,
        uvm_infos=6,
        sim_time_ns=1525.0,
        commit_hash="abc1234",
        run_date="2026-04-16",
        coverage_pct=41.7,
        log_lines=[],
    )


def _make_failing_sim_result(log_path: str = "xsim.log") -> SimResult:
    """Return a SimResult representing a failing simulation run."""
    return SimResult(
        log_path=log_path,
        passed=False,
        uvm_errors=3,
        uvm_fatals=0,
        uvm_warnings=1,
        uvm_infos=6,
        sim_time_ns=500.0,
        commit_hash="dead000",
        run_date="2026-04-16",
        coverage_pct=0.0,
        log_lines=[],
    )


# ── xsim.log helpers ─────────────────────────────────────────────────────────

_UVM_SUMMARY_CLEAN = """\
UVM_INFO @ 0: reporter [RNTST] Running test foo_test...
UVM_INFO C:/tools/uvm.sv(100) @ 0: reporter [UVM/RELNOTES]
UVM_INFO C:/tools/uvm.sv(1000) @ 1525000: reporter [UVM/REPORT/SERVER]
--- UVM Report Summary ---

** Report counts by severity
UVM_INFO :    6
UVM_WARNING :    0
UVM_ERROR :    0
UVM_FATAL :    0
** Report counts by id
[RNTST]     1

$finish called at time : 1525 ns : File "C:/tools/uvm.sv" Line 100
exit
INFO: [Common 17-206] Exiting xsim at Mon Apr 13 18:52:32 2026...
"""

_UVM_SUMMARY_FAILING = """\
UVM_INFO @ 0: reporter [RNTST] Running test foo_test...
UVM_ERROR C:/tests/foo_test.sv(50) @ 100000: uvm_test_top [ERR] Check failed
UVM_ERROR C:/tests/foo_test.sv(51) @ 110000: uvm_test_top [ERR] Check failed
UVM_ERROR C:/tests/foo_test.sv(52) @ 120000: uvm_test_top [ERR] Check failed
UVM_INFO C:/tools/uvm.sv(1000) @ 1525000: reporter [UVM/REPORT/SERVER]
--- UVM Report Summary ---

** Report counts by severity
UVM_INFO :    6
UVM_WARNING :    0
UVM_ERROR :    3
UVM_FATAL :    0

$finish called at time : 500 ns : File "C:/tools/uvm.sv" Line 100
exit
"""


# ── Tests: parse_xsim_log ─────────────────────────────────────────────────────

def test_parse_xsim_log_clean_run(tmp_path):
    """A log with UVM_ERROR=0 and UVM_FATAL=0 should yield passed=True."""
    log_file = tmp_path / "xsim.log"
    log_file.write_text(_UVM_SUMMARY_CLEAN, encoding="utf-8")

    result = parse_xsim_log(str(log_file))

    assert result.passed is True
    assert result.uvm_errors == 0
    assert result.uvm_fatals == 0
    assert result.uvm_warnings == 0
    assert result.uvm_infos == 6


def test_parse_xsim_log_failing_run(tmp_path):
    """A log with UVM_ERROR=3 should yield passed=False and uvm_errors=3."""
    log_file = tmp_path / "xsim.log"
    log_file.write_text(_UVM_SUMMARY_FAILING, encoding="utf-8")

    result = parse_xsim_log(str(log_file))

    assert result.passed is False
    assert result.uvm_errors == 3
    assert result.uvm_fatals == 0


def test_parse_xsim_log_extracts_sim_time(tmp_path):
    """Simulation end time is extracted from the $finish line."""
    log_file = tmp_path / "xsim.log"
    log_file.write_text(_UVM_SUMMARY_CLEAN, encoding="utf-8")

    result = parse_xsim_log(str(log_file))

    assert result.sim_time_ns == pytest.approx(1525.0)


def test_parse_xsim_log_returns_last_50_lines(tmp_path):
    """log_lines contains at most 50 lines."""
    body = "\n".join(f"line {i}" for i in range(200))
    log_file = tmp_path / "xsim.log"
    log_file.write_text(body, encoding="utf-8")

    result = parse_xsim_log(str(log_file))

    assert len(result.log_lines) <= 50


def test_parse_xsim_log_run_date_is_today(tmp_path):
    """run_date is set to today's ISO date."""
    from datetime import date
    log_file = tmp_path / "xsim.log"
    log_file.write_text(_UVM_SUMMARY_CLEAN, encoding="utf-8")

    result = parse_xsim_log(str(log_file))

    assert result.run_date == date.today().isoformat()


def test_parse_xsim_log_fatal_fails(tmp_path):
    """A log with UVM_FATAL=1 should yield passed=False."""
    log = (
        "** Report counts by severity\n"
        "UVM_INFO :    2\n"
        "UVM_WARNING :    0\n"
        "UVM_ERROR :    0\n"
        "UVM_FATAL :    1\n"
    )
    log_file = tmp_path / "xsim.log"
    log_file.write_text(log, encoding="utf-8")

    result = parse_xsim_log(str(log_file))

    assert result.passed is False
    assert result.uvm_fatals == 1


def test_parse_xsim_log_coverage_pct(tmp_path):
    """Coverage percentage is extracted from [COV] reporter lines."""
    log = (
        _UVM_SUMMARY_CLEAN
        + "UVM_INFO ... @ 1525000: uvm_test_top.env_h.cov [COV] "
        "axi_transaction_cg coverage: 41.7%\n"
    )
    log_file = tmp_path / "xsim.log"
    log_file.write_text(log, encoding="utf-8")

    result = parse_xsim_log(str(log_file))

    assert result.coverage_pct == pytest.approx(41.7)


# ── Tests: write_vpr_results ─────────────────────────────────────────────────

def test_write_vpr_results_updates_rtl_columns(tmp_path):
    """Passing sim result writes PASS to all non-WAIVED rows; WAIVED rows untouched."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    data_rows = [
        _make_data_row("TEST-001", disposition="GENERATED", rtl_status="NOT_RUN"),
        _make_data_row("TEST-002", disposition="GENERATED", rtl_status="NOT_RUN"),
        _make_data_row("TEST-003", disposition="WAIVED",    rtl_status="NOT_RUN"),
    ]
    _make_vpr_workbook(data_rows, vpr_path)

    sim = _make_passing_sim_result()
    rows_updated = write_vpr_results(vpr_path, sim)

    assert rows_updated == 2  # WAIVED row excluded

    # Re-open and inspect
    wb = openpyxl.load_workbook(vpr_path, read_only=True, data_only=True)
    ws = wb["VPR"]
    rows = list(ws.iter_rows(min_row=4))
    wb.close()

    # Row 4 → TEST-001 GENERATED → PASS
    assert rows[0][18].value == "PASS"
    assert rows[0][19].value == "2026-04-16"
    assert rows[0][20].value == "abc1234"

    # Row 5 → TEST-002 GENERATED → PASS
    assert rows[1][18].value == "PASS"

    # Row 6 → TEST-003 WAIVED → still NOT_RUN (untouched)
    assert rows[2][18].value == "NOT_RUN"


def test_write_vpr_results_never_downgrades(tmp_path):
    """A row that already has RTL_Status='PASS' must not be set to FAIL."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    data_rows = [
        _make_data_row("KEEP-001", disposition="GENERATED", rtl_status="PASS"),
        _make_data_row("NEW-001",  disposition="GENERATED", rtl_status="NOT_RUN"),
    ]
    _make_vpr_workbook(data_rows, vpr_path)

    sim = _make_failing_sim_result()
    rows_updated = write_vpr_results(vpr_path, sim)

    # Only NEW-001 should be updated (KEEP-001 is already PASS, must not downgrade)
    assert rows_updated == 1

    wb = openpyxl.load_workbook(vpr_path, read_only=True, data_only=True)
    ws = wb["VPR"]
    rows = list(ws.iter_rows(min_row=4))
    wb.close()

    assert rows[0][18].value == "PASS"   # KEEP-001 unchanged
    assert rows[1][18].value == "FAIL"   # NEW-001 updated


def test_write_vpr_results_req_ids_filter(tmp_path):
    """When req_ids is provided, only the specified rows are updated."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    data_rows = [
        _make_data_row("FILT-001", rtl_status="NOT_RUN"),
        _make_data_row("FILT-002", rtl_status="NOT_RUN"),
        _make_data_row("FILT-003", rtl_status="NOT_RUN"),
    ]
    _make_vpr_workbook(data_rows, vpr_path)

    sim = _make_passing_sim_result()
    rows_updated = write_vpr_results(vpr_path, sim, req_ids=["FILT-001", "FILT-003"])

    assert rows_updated == 2

    wb = openpyxl.load_workbook(vpr_path, read_only=True, data_only=True)
    ws = wb["VPR"]
    rows = list(ws.iter_rows(min_row=4))
    wb.close()

    assert rows[0][18].value == "PASS"      # FILT-001 updated
    assert rows[1][18].value == "NOT_RUN"   # FILT-002 skipped
    assert rows[2][18].value == "PASS"      # FILT-003 updated


def test_collect_results_skips_waived(tmp_path):
    """WAIVED rows are never written to, even when req_ids filter is None."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    data_rows = [
        _make_data_row("WAV-001", disposition="WAIVED",    rtl_status="NOT_RUN"),
        _make_data_row("WAV-002", disposition="WAIVED",    rtl_status="NOT_RUN"),
        _make_data_row("GEN-001", disposition="GENERATED", rtl_status="NOT_RUN"),
    ]
    _make_vpr_workbook(data_rows, vpr_path)

    sim = _make_passing_sim_result()
    rows_updated = write_vpr_results(vpr_path, sim, req_ids=None)

    assert rows_updated == 1  # only GEN-001

    wb = openpyxl.load_workbook(vpr_path, read_only=True, data_only=True)
    ws = wb["VPR"]
    rows = list(ws.iter_rows(min_row=4))
    wb.close()

    assert rows[0][18].value == "NOT_RUN"  # WAV-001 untouched
    assert rows[1][18].value == "NOT_RUN"  # WAV-002 untouched
    assert rows[2][18].value == "PASS"     # GEN-001 updated


def test_write_vpr_results_overall_status_not_touched(tmp_path):
    """Overall_Status column (col 27) must never be written by pssgen."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    data_rows = [
        _make_data_row("CHK-001", disposition="GENERATED", rtl_status="NOT_RUN"),
    ]
    # Set a sentinel value in Overall_Status column
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VPR"
    ws.append(["Group headers"] + [None] * (_N_COLS - 1))
    ws.append(_make_header_row())
    ws.append(["[BLOCK-FAM-NNN]"] + [None] * (_N_COLS - 1))
    row = _make_data_row("CHK-001")
    row[27] = "SENTINEL_VALUE"   # Overall_Status sentinel
    ws.append(row + [None] * (_N_COLS - len(row)))
    ws_cov = wb.create_sheet("Coverage_Goals")
    ws_cov.append(["ID", "Name", "Description", "Stimulus_Strategy",
                   "Boundary_Values", "Linked_Requirements", "Status",
                   "Coverage_Type", "Notes"])
    wb.save(vpr_path)

    sim = _make_passing_sim_result()
    write_vpr_results(vpr_path, sim)

    wb2 = openpyxl.load_workbook(vpr_path, read_only=True, data_only=True)
    ws2 = wb2["VPR"]
    data_row = list(ws2.iter_rows(min_row=4))[0]
    wb2.close()

    # Overall_Status must be untouched
    assert data_row[27].value == "SENTINEL_VALUE"
    # RTL_Status must have been written
    assert data_row[18].value == "PASS"


# ── Tests: generate_gap_report_json ──────────────────────────────────────────

def test_generate_gap_report_json_structure(tmp_path):
    """Output JSON must contain the expected top-level keys and summary fields."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    data_rows = [
        _make_data_row("GJSON-001", disposition="GENERATED", rtl_status="PASS"),
        _make_data_row("GJSON-002", disposition="GENERATED", rtl_status="NOT_RUN"),
        _make_data_row("GJSON-003", disposition="WAIVED",    rtl_status="NOT_RUN"),
    ]
    _make_vpr_workbook(data_rows, vpr_path)

    sim = _make_passing_sim_result()
    out_path = str(tmp_path / "gap_report.json")
    generate_gap_report_json(vpr_path, sim, out_path)

    assert os.path.isfile(out_path)
    with open(out_path, encoding="utf-8") as fh:
        report = json.load(fh)

    # Top-level keys
    for key in ("generated", "commit", "sim_result", "summary",
                "family_summary", "requirements"):
        assert key in report, f"Missing top-level key: {key}"

    # Summary keys
    for key in ("total", "waived", "passing", "not_run"):
        assert key in report["summary"], f"Missing summary key: {key}"

    # Requirements list
    assert isinstance(report["requirements"], list)
    req_ids_in_report = {r["req_id"] for r in report["requirements"]}
    assert "GJSON-001" in req_ids_in_report
    assert "GJSON-002" in req_ids_in_report
    assert "GJSON-003" in req_ids_in_report

    # Summary counts (3 total: 2 GENERATED, 1 WAIVED)
    assert report["summary"]["total"] == 3
    assert report["summary"]["waived"] == 1

    # family_summary structure
    fs = report["family_summary"]
    assert isinstance(fs, dict)
    for fam_entry in fs.values():
        for key in ("total", "passing", "failing", "waived", "not_run"):
            assert key in fam_entry, f"Missing family_summary key: {key}"
        assert (
            fam_entry["passing"] + fam_entry["failing"]
            + fam_entry["waived"] + fam_entry["not_run"]
        ) == fam_entry["total"]


def test_generate_gap_report_json_sim_result_block(tmp_path):
    """sim_result block in JSON reflects the SimResult fields."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    _make_vpr_workbook([], vpr_path)

    sim = _make_passing_sim_result()
    out_path = str(tmp_path / "gap_report.json")
    generate_gap_report_json(vplan_path=vpr_path, sim_result=sim, out_path=out_path)

    with open(out_path, encoding="utf-8") as fh:
        report = json.load(fh)

    sr = report["sim_result"]
    assert sr["passed"] is True
    assert sr["uvm_errors"] == 0
    assert sr["uvm_fatals"] == 0
    assert sr["coverage_pct"] == pytest.approx(41.7)


def test_generate_gap_report_json_req_fields(tmp_path):
    """Each requirement entry must have req_id, family, disposition, rtl_status."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    data_rows = [
        _make_data_row("RF-001", disposition="GENERATED"),
    ]
    _make_vpr_workbook(data_rows, vpr_path)

    sim = _make_passing_sim_result()
    out_path = str(tmp_path / "gap_report.json")
    generate_gap_report_json(vpr_path, sim, out_path)

    with open(out_path, encoding="utf-8") as fh:
        report = json.load(fh)

    reqs = {r["req_id"]: r for r in report["requirements"]}
    assert "RF-001" in reqs
    entry = reqs["RF-001"]
    for field_name in ("req_id", "family", "disposition", "covered_by",
                       "rtl_status", "overall_status"):
        assert field_name in entry, f"Missing requirement field: {field_name}"


def test_family_summary_counts(tmp_path):
    """family_summary correctly groups pass/fail/waived counts by family."""
    vpr_path = str(tmp_path / "test_vpr.xlsx")
    data_rows = [
        _make_data_row("BR-001", disposition="GENERATED", rtl_status="PASS",    family="BR"),
        _make_data_row("BR-002", disposition="GENERATED", rtl_status="PASS",    family="BR"),
        _make_data_row("BR-003", disposition="GENERATED", rtl_status="PASS",    family="BR"),
        _make_data_row("BR-004", disposition="WAIVED",    rtl_status="NOT_RUN", family="BR"),
        _make_data_row("FF-001", disposition="GENERATED", rtl_status="FAIL",    family="FF"),
        _make_data_row("FF-002", disposition="GENERATED", rtl_status="FAIL",    family="FF"),
    ]
    _make_vpr_workbook(data_rows, vpr_path)

    sim = _make_passing_sim_result()
    out_path = str(tmp_path / "gap_report.json")
    generate_gap_report_json(vpr_path, sim, out_path)

    with open(out_path, encoding="utf-8") as fh:
        report = json.load(fh)

    fs = report["family_summary"]
    assert fs["BR"]["total"]   == 4
    assert fs["BR"]["passing"] == 3
    assert fs["BR"]["waived"]  == 1
    assert fs["BR"]["failing"] == 0
    assert fs["BR"]["not_run"] == 0
    assert fs["FF"]["total"]   == 2
    assert fs["FF"]["failing"] == 2
    assert fs["FF"]["passing"] == 0
    assert fs["FF"]["waived"]  == 0
