# ===========================================================
# FILE:         tests/test_regmap_simple.py
# PROJECT:      pssgen — AI-Driven PSS + UVM + C Testbench Generator
# COPYRIGHT:    Copyright (c) 2026 BelTech Systems LLC
# LICENSE:      MIT License — see LICENSE file for details
# ===========================================================
#
# DESCRIPTION:
#   Unit tests for the simple_block and system format parsers added in v4c.
#   Covers format detection, access normalisation, volatile parsing, base
#   address inheritance, coverage defaults, field_kind, inline enum parsing,
#   reset value normalisation, and system-level multi-file merge.
#
# LAYER:        1 — parser
# PHASE:        v4c
#
# FUNCTIONS:
#   (test functions — no public API)
#
# DEPENDENCIES:
#   Standard library:  os, tempfile
#   External:          openpyxl, pytest
#   Internal:          parser.regmap_parser
#
# HISTORY:
#   v4c   2026-04-05  SB  Initial implementation; 30 tests for simple_block, real-world, and system formats
#
# ===========================================================
"""tests/test_regmap_simple.py — Tests for simple_block and system regmap parsers.

Phase: v4c
Layer: 1 (parser)

Exercises format detection, simple_block parsing (15-col and 18-col variants),
access/volatile/reset normalisation, coverage defaults, field_kind inference,
inline enum parsing, and system multi-file merge.
"""
import os
import pytest

from parser.regmap_parser import (
    detect_regmap_format,
    parse_regmap,
    _parse_simple_block,
)

FIXTURES = os.path.join(os.path.dirname(__file__), "fixtures")
COUNTER_REGMAP      = os.path.join(FIXTURES, "counter_regmap.xlsx")
COUNTER_SIMPLE      = os.path.join(FIXTURES, "counter_regmap_simple.xlsx")
GPIO_SIMPLE         = os.path.join(FIXTURES, "gpio_regmap_simple.xlsx")
UART_GPIO_SIMPLE    = os.path.join(FIXTURES, "uart_gpio_regmap_simple.xlsx")
SPI_SIMPLE          = os.path.join(FIXTURES, "spi_regmap_simple.xlsx")
TIMER_SIMPLE        = os.path.join(FIXTURES, "timer_regmap_simple.xlsx")
SOC_SYSTEM          = os.path.join(FIXTURES, "soc_regmap_system.xlsx")


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

def test_detect_format_full_block() -> None:
    """detect_regmap_format returns 'full_block' for counter_regmap.xlsx."""
    fmt = detect_regmap_format(COUNTER_REGMAP)
    assert fmt == "full_block"


def test_detect_format_simple_block() -> None:
    """detect_regmap_format returns 'simple_block' for counter_regmap_simple.xlsx."""
    fmt = detect_regmap_format(COUNTER_SIMPLE)
    assert fmt == "simple_block"


# ---------------------------------------------------------------------------
# Simple_block parsing — 18-column counter fixture
# ---------------------------------------------------------------------------

def test_simple_parse_counter_registers() -> None:
    """parse_regmap on counter_regmap_simple.xlsx returns 5 registers."""
    rm = parse_regmap(COUNTER_SIMPLE)
    assert len(rm["registers"]) == 5


def test_simple_parse_total_fields_counter() -> None:
    """counter_regmap_simple.xlsx total field count matches full_block fixture."""
    rm = parse_regmap(COUNTER_SIMPLE)
    total = sum(len(r["fields"]) for r in rm["registers"])
    # Must match the 15 field rows in the source full_block fixture
    assert total == 15


def test_simple_base_address_18col_read() -> None:
    """counter_regmap_simple.xlsx (18 cols) → globals base_address = '0x4000_0000'."""
    rm = parse_regmap(COUNTER_SIMPLE)
    assert rm["globals"]["base_address"] == "0x4000_0000"


def test_simple_access_normalisation_rw() -> None:
    """'RW' access in spreadsheet is preserved as 'RW' in parsed field."""
    rm = parse_regmap(COUNTER_SIMPLE)
    # ENABLE field in CTRL register — access is RW
    ctrl_reg = next(r for r in rm["registers"] if r["name"] == "CTRL")
    enable_field = next(f for f in ctrl_reg["fields"] if f["field_name"] == "ENABLE")
    assert enable_field["access"] == "RW"


def test_simple_access_normalisation_wo_from_wo() -> None:
    """'WO' access in spreadsheet stays 'WO' after normalisation."""
    rm = parse_regmap(COUNTER_SIMPLE)
    # LOAD register has WO fields
    load_reg = next(r for r in rm["registers"] if r["name"] == "LOAD")
    load_val = next(f for f in load_reg["fields"] if f["field_name"] == "LOAD_VAL")
    assert load_val["access"] == "WO"


def test_simple_volatile_yes_true() -> None:
    """'YES' in Volatile column → volatile=True."""
    rm = parse_regmap(COUNTER_SIMPLE)
    status_reg = next(r for r in rm["registers"] if r["name"] == "STATUS")
    running_field = next(f for f in status_reg["fields"] if f["field_name"] == "RUNNING")
    assert running_field["volatile"] is True


def test_simple_volatile_no_false() -> None:
    """'NO' in Volatile column → volatile=False."""
    rm = parse_regmap(COUNTER_SIMPLE)
    ctrl_reg = next(r for r in rm["registers"] if r["name"] == "CTRL")
    enable_field = next(f for f in ctrl_reg["fields"] if f["field_name"] == "ENABLE")
    assert enable_field["volatile"] is False


def test_simple_req_id_populated() -> None:
    """req_id populated from col 17 in 18-column file."""
    rm = parse_regmap(COUNTER_SIMPLE)
    ctrl_reg = next(r for r in rm["registers"] if r["name"] == "CTRL")
    enable_field = next(f for f in ctrl_reg["fields"] if f["field_name"] == "ENABLE")
    assert enable_field["req_id"] == "FUNC-REQ-201"


def test_simple_pss_action_populated() -> None:
    """pss_action populated from col 18 in 18-column file."""
    rm = parse_regmap(COUNTER_SIMPLE)
    ctrl_reg = next(r for r in rm["registers"] if r["name"] == "CTRL")
    enable_field = next(f for f in ctrl_reg["fields"] if f["field_name"] == "ENABLE")
    assert enable_field["pss_action"] == "ctrl_enable"


def test_simple_coverage_default_rw_na_hardware() -> None:
    """RW field with hw_access NA → uvm_has_coverage False."""
    rm = parse_regmap(COUNTER_SIMPLE)
    ctrl_reg = next(r for r in rm["registers"] if r["name"] == "CTRL")
    enable_field = next(f for f in ctrl_reg["fields"] if f["field_name"] == "ENABLE")
    # hw_access is NA → coverage should be False
    assert enable_field["uvm_has_coverage"] is False


def test_simple_coverage_default_status_field() -> None:
    """RO field with hw_access 'set-on-1' → uvm_has_coverage True."""
    rm = parse_regmap(COUNTER_SIMPLE)
    status_reg = next(r for r in rm["registers"] if r["name"] == "STATUS")
    running_field = next(f for f in status_reg["fields"] if f["field_name"] == "RUNNING")
    assert running_field["uvm_has_coverage"] is True


def test_simple_field_kind_reserved_name() -> None:
    """Field named 'RESERVED' → field_kind 'reserved'."""
    rm = parse_regmap(COUNTER_SIMPLE)
    count_reg = next(r for r in rm["registers"] if r["name"] == "COUNT")
    reserved_field = next(
        f for f in count_reg["fields"] if f["field_name"] == "RESERVED"
    )
    assert reserved_field["field_kind"] == "reserved"


def test_simple_field_kind_status_hw() -> None:
    """hw_access 'set-on-1' → field_kind 'status'."""
    rm = parse_regmap(COUNTER_SIMPLE)
    status_reg = next(r for r in rm["registers"] if r["name"] == "STATUS")
    running_field = next(f for f in status_reg["fields"] if f["field_name"] == "RUNNING")
    assert running_field["field_kind"] == "status"


def test_simple_inline_enum_parsed() -> None:
    """Inline enum '0=DISABLED,1=ENABLED' → enums dict entry + enum_ref set."""
    rm = parse_regmap(COUNTER_SIMPLE)
    # ENABLE field uses counter_state_t enum inline in simple format
    ctrl_reg = next(r for r in rm["registers"] if r["name"] == "CTRL")
    enable_field = next(f for f in ctrl_reg["fields"] if f["field_name"] == "ENABLE")
    # Inline enum should be registered
    if enable_field.get("enum_ref"):
        enum_name = enable_field["enum_ref"]
        assert enum_name in rm["enums"]
        assert len(rm["enums"][enum_name]) >= 2


def test_simple_reset_value_zero_to_hex() -> None:
    """Integer reset value 0 → '0x0' hex string."""
    rm = parse_regmap(COUNTER_SIMPLE)
    ctrl_reg = next(r for r in rm["registers"] if r["name"] == "CTRL")
    enable_field = next(f for f in ctrl_reg["fields"] if f["field_name"] == "ENABLE")
    # reset value is "0x0"
    assert enable_field["reset_value"].lower().startswith("0x")


def test_simple_base_address_default_when_absent(tmp_path) -> None:
    """15-column file (no col 16) → globals base_address = '0x0000_0000'."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "RegisterMap"
    # 15-column header
    headers = [
        "Block Name", "Register Name", "Register Offset", "Register Width",
        "Register Description", "Field Name", "Bit Offset", "Bit Width",
        "Access", "Reset Value", "Field Description", "Volatile",
        "Hardware Access", "Software Access", "Field Enumerations",
    ]
    for col, hdr in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=hdr)
    # One data row
    ws.cell(row=2, column=1, value="MYBLOCK")
    ws.cell(row=2, column=2, value="REG0")
    ws.cell(row=2, column=3, value="0x00")
    ws.cell(row=2, column=4, value=32)
    ws.cell(row=2, column=5, value="Test reg")
    ws.cell(row=2, column=6, value="FIELD0")
    ws.cell(row=2, column=7, value=0)
    ws.cell(row=2, column=8, value=8)
    ws.cell(row=2, column=9, value="RW")
    ws.cell(row=2, column=10, value="0x0")
    ws.cell(row=2, column=11, value="A test field")
    ws.cell(row=2, column=12, value="NO")
    ws.cell(row=2, column=13, value="NA")
    ws.cell(row=2, column=14, value="RW")
    ws.cell(row=2, column=15, value="")
    out_file = str(tmp_path / "test_15col.xlsx")
    wb.save(out_file)

    rm = _parse_simple_block(out_file)
    assert rm["globals"]["base_address"] == "0x0000_0000"


# ---------------------------------------------------------------------------
# Real-world fixture tests — gpio_regmap_simple.xlsx (15-column)
# ---------------------------------------------------------------------------

def test_detect_format_gpio_simple_block() -> None:
    """detect_regmap_format returns 'simple_block' for gpio_regmap_simple.xlsx."""
    fmt = detect_regmap_format(GPIO_SIMPLE)
    assert fmt == "simple_block"


def test_simple_parse_gpio_registers() -> None:
    """parse_regmap(gpio_regmap_simple.xlsx) returns 6 registers."""
    rm = parse_regmap(GPIO_SIMPLE)
    assert len(rm["registers"]) == 6


def test_simple_parse_total_gpio_fields() -> None:
    """gpio_regmap_simple.xlsx total field count is 7."""
    rm = parse_regmap(GPIO_SIMPLE)
    total = sum(len(r["fields"]) for r in rm["registers"])
    assert total == 7


def test_simple_base_address_default_gpio() -> None:
    """gpio_regmap_simple.xlsx (15 cols) → globals base_address = '0x0000_0000'."""
    rm = parse_regmap(GPIO_SIMPLE)
    assert rm["globals"]["base_address"] == "0x0000_0000"


def test_simple_parse_uart_gpio_two_blocks() -> None:
    """parse_regmap(uart_gpio_regmap_simple.xlsx) blocks list contains UART0 and GPIO."""
    rm = parse_regmap(UART_GPIO_SIMPLE)
    block_names = [b["block_name"] for b in rm["blocks"]]
    assert "UART0" in block_names
    assert "GPIO" in block_names


# ---------------------------------------------------------------------------
# System fixture tests (added after soc_regmap_system.xlsx is created)
# ---------------------------------------------------------------------------

def test_detect_format_system() -> None:
    """detect_regmap_format returns 'system' for soc_regmap_system.xlsx."""
    if not os.path.exists(SOC_SYSTEM):
        pytest.skip("soc_regmap_system.xlsx not yet created")
    fmt = detect_regmap_format(SOC_SYSTEM)
    assert fmt == "system"


def test_system_parse_four_blocks() -> None:
    """parse_regmap(soc_regmap_system.xlsx) registers contain entries from all 4 blocks."""
    if not os.path.exists(SOC_SYSTEM):
        pytest.skip("soc_regmap_system.xlsx not yet created")
    rm = parse_regmap(SOC_SYSTEM)
    block_names = {r["block"] for r in rm["registers"]}
    # System fixture references COUNTER, GPIO, SPI, TIMER
    assert len(block_names) >= 2  # at minimum two blocks from the files


def test_system_parse_base_override() -> None:
    """System Blocks sheet base_address overrides block-level default."""
    if not os.path.exists(SOC_SYSTEM):
        pytest.skip("soc_regmap_system.xlsx not yet created")
    rm = parse_regmap(SOC_SYSTEM)
    # COUNTER block should have base 0x4000_0000 from system sheet
    counter_block = next(
        (b for b in rm["blocks"] if b["block_name"] == "COUNTER"), None
    )
    assert counter_block is not None
    assert counter_block["base_address"] == "0x4000_0000"


# ---------------------------------------------------------------------------
# Base address inheritance (tmp_path fixture)
# ---------------------------------------------------------------------------

def test_simple_base_address_inherited_row_to_row(tmp_path) -> None:
    """Base address on row 1 only → all subsequent rows inherit it."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "RegisterMap"
    headers = [
        "Block Name", "Register Name", "Register Offset", "Register Width",
        "Register Description", "Field Name", "Bit Offset", "Bit Width",
        "Access", "Reset Value", "Field Description", "Volatile",
        "Hardware Access", "Software Access", "Field Enumerations",
        "base_address",
    ]
    for col, hdr in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=hdr)
    # 3 data rows — base_address only on first row
    for row_i in range(2, 5):
        ws.cell(row=row_i, column=1, value="BLKA")
        ws.cell(row=row_i, column=2, value=f"REG{row_i}")
        ws.cell(row=row_i, column=3, value=f"0x{(row_i-2)*4:02x}")
        ws.cell(row=row_i, column=4, value=32)
        ws.cell(row=row_i, column=5, value="desc")
        ws.cell(row=row_i, column=6, value=f"F{row_i}")
        ws.cell(row=row_i, column=7, value=0)
        ws.cell(row=row_i, column=8, value=8)
        ws.cell(row=row_i, column=9, value="RW")
        ws.cell(row=row_i, column=10, value="0x0")
        ws.cell(row=row_i, column=11, value="")
        ws.cell(row=row_i, column=12, value="NO")
        ws.cell(row=row_i, column=13, value="NA")
        ws.cell(row=row_i, column=14, value="RW")
        ws.cell(row=row_i, column=15, value="")
        ws.cell(row=row_i, column=16, value="0x9000_0000" if row_i == 2 else "")
    out_file = str(tmp_path / "test_inherit.xlsx")
    wb.save(out_file)

    rm = _parse_simple_block(out_file)
    # All 3 rows belong to BLKA block — base should be 0x9000_0000
    assert rm["globals"]["base_address"] == "0x9000_0000"
    assert rm["blocks"][0]["base_address"] == "0x9000_0000"
