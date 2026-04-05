"""
Generates tests/fixtures/counter_regmap_simple.xlsx
18-column simple_block format derived from the full counter_regmap.xlsx.
Run from the project root with the venv active:
    python tests/fixtures/make_counter_regmap_simple.py
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from parser.regmap_parser import _parse_xlsx

# ── Palette ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
LIGHT_GRAY = "F2F2F2"
WHITE      = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
GRAY_FILL   = PatternFill("solid", fgColor=LIGHT_GRAY)
WHITE_FILL  = PatternFill("solid", fgColor=WHITE)

HEADERS_18 = [
    "Block Name", "Register Name", "Register Offset",
    "Register Width", "Register Description", "Field Name",
    "Bit Offset", "Bit Width", "Access", "Reset Value",
    "Field Description", "Volatile", "Hardware Access",
    "Software Access", "Field Enumerations",
    "base_address", "req_id", "pss_action",
]


def hdr_font():
    return Font(name="Arial", bold=True, color=WHITE, size=10)


def body_font():
    return Font(name="Arial", size=10)


def make_header_row(ws, headers):
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def write_row(ws, row_num, values, fill=None):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill


def auto_width(ws, min_width=12):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 45))
        ws.column_dimensions[col_letter].width = max_len + 2


def format_enum(enum_ref: str, enums: dict) -> str:
    """Format enum values as '0=SYM,1=SYM' string."""
    if not enum_ref or enum_ref not in enums:
        return ""
    parts = []
    for entry in enums[enum_ref]:
        parts.append(f"{entry['value']}={entry['symbol']}")
    return ",".join(parts)


def main():
    fixtures_dir = os.path.dirname(__file__)
    source_xlsx = os.path.join(fixtures_dir, "counter_regmap.xlsx")
    out_xlsx = os.path.join(fixtures_dir, "counter_regmap_simple.xlsx")

    # Parse the full_block format to get normalised data
    regmap = _parse_xlsx(source_xlsx)
    globals_dict = regmap["globals"]
    registers = regmap["registers"]
    enums = regmap["enums"]

    base_address = globals_dict.get("base_address", "0x0000_0000") or "0x0000_0000"

    wb = Workbook()
    ws = wb.active
    ws.title = "RegisterMap"
    ws.freeze_panes = "A2"

    make_header_row(ws, HEADERS_18)

    # Alternating fill by register band — track register index
    fills = [WHITE_FILL, GRAY_FILL]
    row_num = 2
    first_field_row_of_block = True  # only write base_address on first field row

    for reg_idx, reg in enumerate(registers):
        fill = fills[reg_idx % 2]
        is_first_reg = (reg_idx == 0)
        for field in reg.get("fields", []):
            enum_str = format_enum(field.get("enum_ref", ""), enums)
            volatile_str = "YES" if field.get("volatile") else "NO"

            # base_address: only on the very first field row of the block
            if is_first_reg and first_field_row_of_block:
                base_addr_cell = base_address
                first_field_row_of_block = False
            else:
                base_addr_cell = ""

            row_data = [
                reg["block"],                           # 0 Block Name
                reg["name"],                            # 1 Register Name
                reg["offset"],                          # 2 Register Offset
                reg["width"],                           # 3 Register Width
                reg["description"],                     # 4 Register Description
                field["field_name"],                    # 5 Field Name
                field["bit_offset"],                    # 6 Bit Offset
                field["bit_width"],                     # 7 Bit Width
                field["access"],                        # 8 Access
                field["reset_value"],                   # 9 Reset Value
                field["description"],                   # 10 Field Description
                volatile_str,                           # 11 Volatile
                field["hw_access"],                     # 12 Hardware Access
                field["sw_access"],                     # 13 Software Access
                enum_str,                               # 14 Field Enumerations
                base_addr_cell,                         # 15 base_address
                field.get("req_id") or "",              # 16 req_id
                field.get("pss_action") or "",          # 17 pss_action
            ]
            write_row(ws, row_num, row_data, fill=fill)
            row_num += 1

    auto_width(ws)
    wb.save(out_xlsx)
    print(f"Saved: {out_xlsx}")


if __name__ == "__main__":
    main()
