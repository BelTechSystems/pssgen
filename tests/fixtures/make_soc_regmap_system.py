"""
Generates tests/fixtures/soc_regmap_system.xlsx
Two-sheet system spreadsheet referencing four block files.
Run from the project root with the venv active:
    python tests/fixtures/make_soc_regmap_system.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DARK_BLUE  = "1F3864"
LIGHT_GRAY = "F2F2F2"
WHITE      = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
GRAY_FILL   = PatternFill("solid", fgColor=LIGHT_GRAY)
WHITE_FILL  = PatternFill("solid", fgColor=WHITE)


def hdr_font():
    return Font(name="Arial", bold=True, color=WHITE, size=10)


def body_font():
    return Font(name="Arial", size=10)


def make_header_row(ws, headers):
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def write_row(ws, row_num, values, fill=None):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill


def auto_width(ws, min_width=12):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2


def main():
    wb = Workbook()

    # --- Sheet 1: System ---
    ws_sys = wb.active
    ws_sys.title = "System"
    ws_sys.freeze_panes = "A2"
    make_header_row(ws_sys, ["Key", "Value"])

    sys_rows = [
        ("project_name", "peripheral_soc"),
        ("data_width",   "32"),
        ("endianness",   "Little"),
        ("bus_type",     "AXI4-Lite"),
    ]
    fills = [WHITE_FILL, GRAY_FILL]
    for r_idx, (key, val) in enumerate(sys_rows, start=2):
        write_row(ws_sys, r_idx, [key, val], fill=fills[r_idx % 2])
    auto_width(ws_sys)

    # --- Sheet 2: Blocks ---
    ws_blk = wb.create_sheet("Blocks")
    ws_blk.freeze_panes = "A2"
    make_header_row(ws_blk, [
        "block_name", "spreadsheet_file", "base_address", "description"
    ])

    # Only single-block files to avoid duplicate GPIO
    block_rows = [
        ("COUNTER", "counter_regmap_simple.xlsx", "0x4000_0000", "Up/down counter"),
        ("GPIO",    "gpio_regmap_simple.xlsx",    "0x4001_0000", "AXI GPIO module"),
        ("SPI",     "spi_regmap_simple.xlsx",     "0x4002_0000", "SPI master"),
        ("TIMER",   "timer_regmap_simple.xlsx",   "0x4003_0000", "Timer"),
    ]
    for r_idx, row_data in enumerate(block_rows, start=2):
        write_row(ws_blk, r_idx, list(row_data), fill=fills[r_idx % 2])
    auto_width(ws_blk)

    out_path = os.path.join(os.path.dirname(__file__), "soc_regmap_system.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
