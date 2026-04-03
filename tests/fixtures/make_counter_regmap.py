"""
Generates tests/fixtures/counter_regmap.xlsx
Gold reference register map for the 8-bit up/down counter.
Run from the project root with the venv active:
    python tests/fixtures/make_counter_regmap.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
LIGHT_GRAY = "F2F2F2"
WHITE      = "FFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
GRAY_FILL   = PatternFill("solid", fgColor=LIGHT_GRAY)
WHITE_FILL  = PatternFill("solid", fgColor=WHITE)

def hdr_font():
    return Font(name="Arial", bold=True, color=WHITE, size=10)

def body_font():
    return Font(name="Arial", size=10)

def make_header_row(ws, row_num, headers):
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col, value=hdr)
        cell.fill = HEADER_FILL
        cell.font = hdr_font()
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

def write_row(ws, row_num, values, fill=None):
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = body_font()
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill

def auto_width(ws, min_width=12):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 45))
        ws.column_dimensions[col_letter].width = max_len + 2

# ── SHEET: Globals ────────────────────────────────────────────────────────────
def build_globals(wb):
    ws = wb.active
    ws.title = "Globals"

    make_header_row(ws, 1, ["Key", "Value", "Description"])
    ws.freeze_panes = "A2"

    rows = [
        ("project_name",    "up_down_counter",   "Project or IP block name"),
        ("revision",        "1.0",               "Semantic version of this register spec"),
        ("author",          "BelTech Systems LLC","Owner/author"),
        ("date",            "2026-03-28",         "Date of this revision"),
        ("data_width_bits", "32",                 "Bus data width: 8, 16, 32, or 64"),
        ("endianness",      "Little",             "Little or Big"),
        ("bus_type",        "AXI4-Lite",          "AXI4-Lite, APB, Wishbone, or custom"),
        ("base_address",    "0x4000_0000",        "Base address of this IP block"),
        ("address_stride",  "0x04",               "Default register stride in bytes (hex)"),
    ]
    for r_idx, (key, val, desc) in enumerate(rows, start=2):
        write_row(ws, r_idx, [key, val, desc])

    auto_width(ws)
    ws.column_dimensions["C"].width = 40

# ── SHEET: Blocks ─────────────────────────────────────────────────────────────
def build_blocks(wb):
    ws = wb.create_sheet("Blocks")

    make_header_row(ws, 1, [
        "block_name", "base_address", "data_width_bits",
        "reset_domain", "clock_domain", "description"
    ])
    ws.freeze_panes = "A2"

    write_row(ws, 2, [
        "COUNTER", "0x4000_0000", "32", "RST_N", "CLK",
        "8-bit up/down counter with AXI4-Lite register interface"
    ])

    auto_width(ws)

# ── SHEET: RegisterMap ────────────────────────────────────────────────────────
# 20 columns matching the template exactly.
REGMAP_HEADERS = [
    "block_name", "reg_name", "reg_description", "address_offset",
    "reg_width_bits", "field_name", "bit_offset", "bit_width",
    "access", "reset_value", "volatile", "hw_access", "sw_access",
    "field_kind", "enum_ref", "uvm_has_coverage", "req_id",
    "pss_action", "hdl_path", "description",
]

# Each tuple: 20 fields in column order.
# Register bands: CTRL=white, STATUS=gray, COUNT=white, LOAD=gray, INT_CTRL=white
REGMAP_ROWS = [
    # ── Register 1: CTRL ─────────────────────────────────────────────────────
    ("COUNTER","CTRL","Control Register","0x00","32",
     "ENABLE","0","1","RW","0x0","NO","NA","RW",
     "normal","counter_state_t","YES","FUNC-REQ-201","ctrl_enable",
     "","Enable counter operation. 0=disabled 1=enabled."),

    ("COUNTER","CTRL","Control Register","0x00","32",
     "UP_DOWN","1","1","RW","0x1","NO","NA","RW",
     "normal","counter_dir_t","YES","FUNC-REQ-202","ctrl_direction",
     "","Count direction. 0=count down 1=count up."),

    ("COUNTER","CTRL","Control Register","0x00","32",
     "AUTO_RELOAD","2","1","RW","0x0","NO","NA","RW",
     "normal","","YES","FUNC-REQ-203","",
     "","Auto-reload on rollover. 0=stop 1=reload."),

    # ── Register 2: STATUS ────────────────────────────────────────────────────
    ("COUNTER","STATUS","Status Register","0x04","32",
     "RUNNING","0","1","RO","0x0","YES","set-on-1","RO",
     "status","","YES","FUNC-REQ-204","",
     "","Counter is running. Set by HW when ENABLE=1."),

    ("COUNTER","STATUS","Status Register","0x04","32",
     "ROLLOVER","1","1","RO","0x0","YES","set-on-1","RO",
     "status","","YES","FUNC-REQ-205","",
     "","Rollover event occurred since last read."),

    ("COUNTER","STATUS","Status Register","0x04","32",
     "AT_MAX","2","1","RO","0x0","YES","set-on-1","RO",
     "status","","YES","FUNC-REQ-206","",
     "","Count is at maximum value (255)."),

    ("COUNTER","STATUS","Status Register","0x04","32",
     "AT_MIN","3","1","RO","0x0","YES","set-on-1","RO",
     "status","","YES","FUNC-REQ-207","",
     "","Count is at minimum value (0)."),

    # ── Register 3: COUNT ─────────────────────────────────────────────────────
    ("COUNTER","COUNT","Count Value Register","0x08","32",
     "COUNT_VAL","0","8","RO","0x0","YES","set-by-input","RO",
     "status","","YES","FUNC-REQ-208","read_count",
     "dut.counter.count_reg","Current count value. Updated each clock cycle."),

    ("COUNTER","COUNT","Count Value Register","0x08","32",
     "RESERVED","8","24","NA","0x0","NO","NA","NA",
     "reserved","","NO","",
     "","","Reserved. Read as zero. Do not write."),

    # ── Register 4: LOAD ──────────────────────────────────────────────────────
    ("COUNTER","LOAD","Count Load Register","0x0C","32",
     "LOAD_VAL","0","8","WO","0x0","NO","NA","WO",
     "normal","","NO","FUNC-REQ-209","load_count",
     "","Write to preload COUNT_VAL. Takes effect on next clock edge when ENABLE=1."),

    ("COUNTER","LOAD","Count Load Register","0x0C","32",
     "RESERVED","8","24","NA","0x0","NO","NA","NA",
     "reserved","","NO","",
     "","","Reserved. Read as zero. Do not write."),

    # ── Register 5: INT_CTRL ──────────────────────────────────────────────────
    ("COUNTER","INT_CTRL","Interrupt Control Register","0x10","32",
     "ROLLOVER_EN","0","1","RW","0x0","NO","NA","RW",
     "normal","","YES","FUNC-REQ-210","",
     "","Enable interrupt on rollover event."),

    ("COUNTER","INT_CTRL","Interrupt Control Register","0x10","32",
     "AT_MAX_EN","1","1","RW","0x0","NO","NA","RW",
     "normal","","YES","FUNC-REQ-211","",
     "","Enable interrupt when count reaches 255."),

    ("COUNTER","INT_CTRL","Interrupt Control Register","0x10","32",
     "AT_MIN_EN","2","1","RW","0x0","NO","NA","RW",
     "normal","","YES","FUNC-REQ-212","",
     "","Enable interrupt when count reaches 0."),

    ("COUNTER","INT_CTRL","Interrupt Control Register","0x10","32",
     "INT_STATUS","4","4","RC","0x0","YES","set-on-1","RC",
     "status","","YES","FUNC-REQ-213","read_int_status",
     "","Interrupt status bits. Read-to-clear. [7]=AT_MIN [6]=AT_MAX [5]=reserved [4]=ROLLOVER"),
]

# Which register index (0-based, by reg_name change) gets which fill
def register_fill_for_row(row_data):
    reg_bands = {"CTRL": WHITE_FILL, "STATUS": GRAY_FILL,
                 "COUNT": WHITE_FILL, "LOAD": GRAY_FILL, "INT_CTRL": WHITE_FILL}
    return reg_bands.get(row_data[1], WHITE_FILL)

def build_registermap(wb):
    ws = wb.create_sheet("RegisterMap")
    make_header_row(ws, 1, REGMAP_HEADERS)
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(REGMAP_ROWS, start=2):
        fill = register_fill_for_row(row)
        write_row(ws, r_idx, list(row), fill=fill)

    auto_width(ws)

# ── SHEET: Enums ──────────────────────────────────────────────────────────────
ENUM_ROWS = [
    ("counter_dir_t",   "0", "COUNT_DOWN", "Count decrements each cycle"),
    ("counter_dir_t",   "1", "COUNT_UP",   "Count increments each cycle"),
    ("counter_state_t", "0", "DISABLED",   "Counter halted"),
    ("counter_state_t", "1", "ENABLED",    "Counter running"),
]

def build_enums(wb):
    ws = wb.create_sheet("Enums")
    make_header_row(ws, 1, ["enum_name", "value", "symbol", "description"])
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(ENUM_ROWS, start=2):
        write_row(ws, r_idx, list(row))

    auto_width(ws)

# ── Assemble ──────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    build_globals(wb)
    build_blocks(wb)
    build_registermap(wb)
    build_enums(wb)

    out_path = os.path.join(os.path.dirname(__file__), "counter_regmap.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
