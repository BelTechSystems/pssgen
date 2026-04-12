"""Generate buffered_axi_lite_uart_vplan.xlsx from template and source data."""
import shutil
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── 1. Copy template ──────────────────────────────────────────────────────────
SRC = 'docs/pssgen_vpr_template.xlsx'
DST = 'ip/buffered_axi_lite_uart/buffered_axi_lite_uart_vplan.xlsx'
shutil.copy(SRC, DST)
wb = openpyxl.load_workbook(DST)

# ── 2. Static data tables ─────────────────────────────────────────────────────

# Section mapping
SECTION = {}
for k in ['UART-PAR-001','UART-PAR-002','UART-PAR-003','UART-PAR-004','UART-PAR-005',
          'UART-PAR-006','UART-PAR-007','UART-PAR-008','UART-PAR-009']:
    SECTION[k] = '§3.4'
for i in range(1,12):
    SECTION[f'UART-IF-{i:03d}'] = '§6.1'
for i in range(12,15):
    SECTION[f'UART-IF-{i:03d}'] = '§6.2'
for k in ['UART-EN-001','UART-EN-002','UART-EN-003','UART-EN-004','UART-EN-005','UART-EN-006']:
    SECTION[k] = '§7.1'
for i in range(1,7):
    SECTION[f'UART-BR-{i:03d}'] = '§7.2'
for i in range(1,10):
    SECTION[f'UART-FF-{i:03d}'] = '§7.3'
for i in range(1,10):
    SECTION[f'UART-FIFO-{i:03d}'] = '§7.4'
for i in range(1,8):
    SECTION[f'UART-TO-{i:03d}'] = '§7.5'
for i in range(1,14):
    SECTION[f'UART-INT-{i:03d}'] = '§7.6'
for i in range(1,5):
    SECTION[f'UART-REG-{i:03d}'] = '§8.1'
for i in range(5,12):
    SECTION[f'UART-REG-{i:03d}'] = '§8.2'
for i in range(12,26):
    SECTION[f'UART-REG-{i:03d}'] = '§8.3'
for i in range(26,29):
    SECTION[f'UART-REG-{i:03d}'] = '§8.4'
for i in range(29,32):
    SECTION[f'UART-REG-{i:03d}'] = '§8.5'
for i in range(32,35):
    SECTION[f'UART-REG-{i:03d}'] = '§8.6'
for i in range(35,37):
    SECTION[f'UART-REG-{i:03d}'] = '§8.7'
SECTION['UART-REG-037'] = '§8.8'
for i in range(38,41):
    SECTION[f'UART-REG-{i:03d}'] = '§8.9'
for i in range(41,44):
    SECTION[f'UART-REG-{i:03d}'] = '§8.10'
for i in range(44,47):
    SECTION[f'UART-REG-{i:03d}'] = '§8.11'
for i in range(47,50):
    SECTION[f'UART-REG-{i:03d}'] = '§8.12'
for i in range(50,53):
    SECTION[f'UART-REG-{i:03d}'] = '§8.13'
for i in range(1,7):
    SECTION[f'UART-RST-{i:03d}'] = '§9'
for i in range(1,11):
    SECTION[f'UART-VER-{i:03d}'] = '§11'

# Type mapping
TYPE = {}
for k in ['UART-PAR-001','UART-PAR-002','UART-PAR-004','UART-PAR-005','UART-PAR-006',
          'UART-PAR-007','UART-PAR-008','UART-PAR-009']:
    TYPE[k] = 'parametric'
TYPE['UART-PAR-003'] = 'build-time'
for fam in ['IF','EN','BR','FF','FIFO','TO','INT','RST']:
    for i in range(1,20):
        TYPE[f'UART-{fam}-{i:03d}'] = 'functional'
for i in range(1,53):
    TYPE[f'UART-REG-{i:03d}'] = 'structural'
for i in range(1,11):
    TYPE[f'UART-VER-{i:03d}'] = 'infrastructure'

# Covered_By mapping
COVERED_BY = {
    'UART-BR-001':'COV-001','UART-BR-002':'COV-001','UART-BR-003':'COV-001',
    'UART-BR-004':'',        # gap demo: blank until after demo; COV-013 closes this gap
    'UART-BR-005':'COV-001','UART-BR-006':'COV-001',
    'UART-EN-001':'COV-016','UART-EN-002':'COV-016','UART-EN-003':'COV-016',
    'UART-EN-004':'COV-016','UART-EN-005':'COV-016','UART-EN-006':'COV-016',
    'UART-FF-001':'COV-017','UART-FF-002':'COV-003','UART-FF-003':'COV-003',
    'UART-FF-004':'COV-003','UART-FF-005':'COV-002','UART-FF-006':'COV-002',
    'UART-FF-007':'COV-002','UART-FF-008':'COV-002','UART-FF-009':'COV-002',
    'UART-FIFO-001':'COV-017','UART-FIFO-002':'COV-005','UART-FIFO-003':'COV-005',
    'UART-FIFO-004':'COV-004','UART-FIFO-005':'COV-004','UART-FIFO-006':'COV-005',
    'UART-FIFO-007':'COV-004, COV-008','UART-FIFO-008':'COV-005','UART-FIFO-009':'COV-004',
    'UART-IF-001':'COV-017','UART-IF-002':'COV-017','UART-IF-003':'COV-017',
    'UART-IF-004':'COV-017','UART-IF-005':'COV-012','UART-IF-006':'COV-010',
    'UART-IF-007':'COV-011','UART-IF-008':'COV-010','UART-IF-009':'COV-011',
    'UART-IF-010':'COV-018','UART-IF-011':'COV-018','UART-IF-012':'COV-017',
    'UART-IF-013':'COV-017','UART-IF-014':'COV-017',
    'UART-INT-001':'COV-007','UART-INT-002':'COV-007, COV-009','UART-INT-003':'COV-007',
    'UART-INT-004':'COV-007','UART-INT-005':'COV-007','UART-INT-006':'COV-007',
    'UART-INT-007':'COV-007','UART-INT-008':'COV-007, COV-009',
    'UART-INT-009':'COV-019','UART-INT-010':'COV-019','UART-INT-011':'COV-019',
    'UART-INT-012':'COV-019','UART-INT-013':'COV-019',
    'UART-PAR-001':'COV-017','UART-PAR-002':'COV-017','UART-PAR-003':'',
    'UART-PAR-004':'COV-001','UART-PAR-005':'COV-001','UART-PAR-006':'COV-006',
    'UART-PAR-007':'COV-001, COV-015','UART-PAR-008':'COV-015','UART-PAR-009':'COV-017',
    'UART-REG-001':'COV-010, COV-011','UART-REG-002':'COV-018','UART-REG-003':'COV-018',
    'UART-REG-004':'COV-015','UART-REG-005':'COV-016','UART-REG-006':'COV-016',
    'UART-REG-007':'COV-016','UART-REG-008':'COV-016','UART-REG-009':'COV-002',
    'UART-REG-010':'COV-003','UART-REG-011':'COV-018','UART-REG-012':'COV-018',
    'UART-REG-013':'COV-015','UART-REG-014':'COV-009','UART-REG-015':'COV-019',
    'UART-REG-016':'COV-005','UART-REG-017':'COV-005','UART-REG-018':'COV-008',
    'UART-REG-019':'COV-004','UART-REG-020':'COV-019','UART-REG-021':'COV-019',
    'UART-REG-022':'COV-019','UART-REG-023':'COV-019','UART-REG-024':'COV-008',
    'UART-REG-025':'COV-018','UART-REG-026':'COV-001','UART-REG-027':'COV-001',
    'UART-REG-028':'COV-013','UART-REG-029':'COV-019','UART-REG-030':'COV-019',
    'UART-REG-031':'COV-015','UART-REG-032':'COV-017, COV-018',
    'UART-REG-033':'COV-005, COV-017','UART-REG-034':'COV-004, COV-017',
    'UART-REG-035':'COV-006','UART-REG-036':'COV-006, COV-015',
    'UART-REG-037':'COV-007, COV-019','UART-REG-038':'COV-018, COV-019',
    'UART-REG-039':'COV-007, COV-019','UART-REG-040':'COV-007, COV-019',
    'UART-REG-041':'COV-019','UART-REG-042':'COV-018, COV-019','UART-REG-043':'COV-019',
    'UART-REG-044':'COV-014','UART-REG-045':'COV-014','UART-REG-046':'COV-014',
    'UART-REG-047':'COV-019','UART-REG-048':'COV-019','UART-REG-049':'COV-018',
    'UART-REG-050':'COV-019','UART-REG-051':'COV-019','UART-REG-052':'COV-018',
    'UART-RST-001':'COV-015','UART-RST-002':'COV-015','UART-RST-003':'COV-015',
    'UART-RST-004':'COV-015','UART-RST-005':'COV-015','UART-RST-006':'COV-015',
    'UART-TO-001':'COV-006','UART-TO-002':'COV-006','UART-TO-003':'COV-006',
    'UART-TO-004':'COV-006','UART-TO-005':'COV-006, COV-009','UART-TO-006':'COV-006',
    'UART-TO-007':'COV-019',
    'UART-VER-001':'','UART-VER-002':'COV-015','UART-VER-003':'COV-014',
    'UART-VER-004':'COV-007','UART-VER-005':'COV-008','UART-VER-006':'COV-012',
    'UART-VER-007':'COV-013','UART-VER-008':'COV-006','UART-VER-009':'COV-002',
    'UART-VER-010':'COV-008',
}

# Stimulus strategy per COV item
STIM_STRAT = {
    'COV-001': 'NCO tuning word sweep at 12 values from minimum (0x00000001) to maximum (0xFFFFFFFF) including all standard baud rates.',
    'COV-002': 'All four parity modes (00/01/10/11) exercised independently in LOOP_EN loopback; inject parity error bit for odd and even modes.',
    'COV-003': 'Both stop bit configurations (1 and 2 stop bits) exercised in LOOP_EN loopback with frame integrity check.',
    'COV-004': 'RX FIFO occupancy driven to 0, RX_THRESH-1, RX_THRESH, RX_THRESH+1, and G_FIFO_DEPTH to verify status flags and overrun discard.',
    'COV-005': 'TX FIFO occupancy driven to 0, TX_THRESH-1, TX_THRESH, TX_THRESH+1, and G_FIFO_DEPTH to verify status flags and write-ignore behavior.',
    'COV-006': 'TIMEOUT_VAL set to 0x0000, 0x0001, G_TIMEOUT_DEFAULT, 0xFFFE, and 0xFFFF; idle receiver until timeout fires for each non-zero value.',
    'COV-007': 'Each of 8 interrupt sources set and cleared independently via directed stimulus; verify IRQ asserts and deasserts correctly for each bit.',
    'COV-008': 'Fill RX FIFO to capacity then transmit additional byte to force overrun; verify overrun byte discarded and FIFO content unchanged.',
    'COV-009': 'Set TIMEOUT_VAL non-zero, fill RX FIFO above RX_THRESH, then idle until timeout fires; verify both INT_STATUS.TIMEOUT and INT_STATUS.RX_THRESH set simultaneously.',
    'COV-010': 'AXI write to valid register offset (OKAY expected) and to undefined offset 0x30 (SLVERR expected); capture BRESP each time.',
    'COV-011': 'AXI read from valid register offset (OKAY expected) and from undefined offset 0x30 (SLVERR expected); capture RRESP each time.',
    'COV-012': 'Issue AXI write with AWVALID first then WVALID, then repeat with WVALID first then AWVALID; verify correct response in both orderings.',
    'COV-013': 'Assert UART_EN, write a new value to BAUD_TUNING, read back and verify original value unchanged; confirms silent-ignore while enabled.',
    'COV-014': 'Write walking-ones (0x00000001 through 0x80000000), all-ones (0xFFFFFFFF), and all-zeros (0x00000000) to SCRATCH; read back and verify each value.',
    'COV-015': 'Assert then deassert reset; read all 12 registers and compare each field to the reset value specified in Table 8-x.',
    'COV-016': 'Deassert UART_EN and verify TX/RX halt; then independently gate TX_EN and RX_EN with UART_EN=1 to verify path isolation; set LOOP_EN and verify loopback active.',
    'COV-017': 'Structural inspection: verify 8-bit data frames transmitted and received, AXI data width=32, address width sufficient, single-beat transactions only.',
    'COV-018': 'Write to each read-only register (STATUS, FIFO_STATUS, INT_STATUS, INT_CLEAR, TX_DATA reads, RX_DATA writes); read back and verify state unchanged.',
    'COV-019': 'Drive TX_THRESH and RX_THRESH interrupt conditions, inject FRAME_ERR via stop-bit corruption, enable TIMEOUT interrupt; verify each INT_STATUS bit controlled by INT_ENABLE gate.',
}

def get_stim(req_id):
    cov = COVERED_BY.get(req_id, '')
    if not cov:
        return ''
    first_cov = cov.split(',')[0].strip()
    return STIM_STRAT.get(first_cov, '')

# ── 3. Parse requirements from .req file ─────────────────────────────────────
def parse_req_file(path):
    reqs = {}
    verif = {}
    current_id = None
    with open(path, encoding='utf-8') as f:
        for line in f:
            m = re.match(r'^\[([A-Z]+-[A-Z]+-\d+)\]\s+(.*)', line)
            if m:
                current_id = m.group(1)
                reqs[current_id] = m.group(2).strip()
            elif current_id and re.match(r'^\s+verification:\s+(.+)', line):
                mv = re.match(r'^\s+verification:\s+(.+)', line)
                v = mv.group(1).strip()
                if v == '(pending review)':
                    v = 'test'
                verif[current_id] = v
    return reqs, verif

REQS, VERIF = parse_req_file('ip/buffered_axi_lite_uart/buffered_axi_lite_uart.req')

REQ_ORDER = (
    [f'UART-PAR-{i:03d}' for i in range(1,10)] +
    [f'UART-IF-{i:03d}' for i in range(1,15)] +
    [f'UART-EN-{i:03d}' for i in range(1,7)] +
    [f'UART-BR-{i:03d}' for i in range(1,7)] +
    [f'UART-FF-{i:03d}' for i in range(1,10)] +
    [f'UART-FIFO-{i:03d}' for i in range(1,10)] +
    [f'UART-TO-{i:03d}' for i in range(1,8)] +
    [f'UART-INT-{i:03d}' for i in range(1,14)] +
    [f'UART-REG-{i:03d}' for i in range(1,53)] +
    [f'UART-RST-{i:03d}' for i in range(1,7)] +
    [f'UART-VER-{i:03d}' for i in range(1,11)]
)
assert len(REQ_ORDER) == 141, f"Expected 141, got {len(REQ_ORDER)}"

def get_family(req_id):
    return req_id.split('-')[1]

WAIVER_RATIONALE = {
    'UART-PAR-003': ('Elaboration-time assertion — $fatal fires before simulation begins. '
                     'Requires a separate negative build test. Deferred to build-system test phase.'),
    'UART-VER-001': ('Verification infrastructure requirement. Describes the UVM RAL model artifact, '
                     'not a testable behavior. Closed by inspection when the RAL model is committed.'),
}
RISK = {
    'UART-PAR-003': 'MEDIUM',
    'UART-VER-001': 'LOW',
}

# ── 4. Populate VPR tab ───────────────────────────────────────────────────────
ws_vpr = wb['VPR']

data_font   = Font(name='Arial', size=9)
grey_font   = Font(name='Arial', size=9, italic=True, color='595959')
grey_fill   = PatternFill('solid', fgColor='F9F9F9')
wrap_align  = Alignment(horizontal='left', vertical='top', wrap_text=True)
ctr_align   = Alignment(horizontal='center', vertical='top', wrap_text=True)

# Ensure example row 3 stays grey/italic
for col in range(1, 31):
    cell = ws_vpr.cell(row=3, column=col)
    cell.font = grey_font
    cell.fill = grey_fill

# Write data rows 4..144
for i, req_id in enumerate(REQ_ORDER):
    row = 4 + i
    family       = get_family(req_id)
    stmt         = REQS.get(req_id, '')
    verif_method = VERIF.get(req_id, 'test')
    covered_by   = COVERED_BY.get(req_id, '')
    stim         = get_stim(req_id)
    section      = SECTION.get(req_id, '')
    req_type     = TYPE.get(req_id, 'functional')
    disposition  = 'WAIVED' if req_id in WAIVER_RATIONALE else 'GENERATED'
    waiver_rat   = WAIVER_RATIONALE.get(req_id, '')
    risk         = RISK.get(req_id, '')

    row_data = [
        req_id,                              # A  Req_ID
        family,                              # B  Family
        req_type,                            # C  Type
        stmt,                                # D  Statement
        'BALU-RS-001',                       # E  Source_Doc
        section,                             # F  Source_Section
        verif_method,                        # G  Verification_Method
        covered_by,                          # H  Covered_By
        stim,                                # I  Stimulus_Strategy
        '',                                  # J  Boundary_Notes
        disposition,                         # K  Disposition
        waiver_rat,                          # L  Waiver_Rationale
        risk,                                # M  Risk_Acceptance
        '',                                  # N  Test_Name
        'vhdl/buffered_axi_lite_uart.vhd',   # O  HDL_Source
        'RTL',                               # P  Sim_Mode
        'NOT_RUN',                           # Q  RTL_Status
        '',                                  # R  RTL_Run_Date
        '',                                  # S  RTL_Commit
        '',                                  # T  RTL_Evidence
        'NOT_RUN',                           # U  Gate_Status
        '',                                  # V  Gate_Evidence
        '',                                  # W  Anomaly_ID
        None,                                # X  Overall_Status (formula below)
        '',                                  # Y  Closure_Author
        '',                                  # Z  Closure_Date
        '',                                  # AA Closure_Notes
        '1.0.0',                             # AB IP_Version
        'ZUBoard 1CG / Basys 3',             # AC Target_Device
        'BALU-RS-001 Rev 0.4',               # AD Spec_Revision
    ]

    for col, val in enumerate(row_data, 1):
        cell = ws_vpr.cell(row=row, column=col, value=val)
        cell.font = data_font
        cell.alignment = wrap_align

    # Overall_Status formula col X (24)
    formula = (f'=IF(K{row}="WAIVED","WAIVED",'
               f'IF(H{row}="","OPEN",'
               f'IF(Q{row}="PASS",'
               f'IF(Y{row}<>"","CLOSED","PASSING"),'
               f'IF(Q{row}="FAIL","FAILING","COVERED"))))')
    cell_x = ws_vpr.cell(row=row, column=24, value=formula)
    cell_x.font = data_font
    cell_x.alignment = ctr_align

print(f"VPR: wrote {len(REQ_ORDER)} data rows (rows 4..{3+len(REQ_ORDER)})")

# ── 5. Coverage_Goals tab ─────────────────────────────────────────────────────
ws_cg = wb['Coverage_Goals']
for row in ws_cg.iter_rows(min_row=3, max_row=max(ws_cg.max_row, 25)):
    for cell in row:
        cell.value = None

cov_data = [
    ('COV-001','BAUD_TUNING',
     'NCO tuning word sweep covering standard baud rates and boundary values. 12 values computed for G_CLK_FREQ_HZ = 100 MHz.',
     STIM_STRAT['COV-001'],
     '0x00000001 (min), standard baud rates 9600-921600, 3/4 Mbaud, 0xFFFFFFFF (max)',
     'UART-BR-001, UART-BR-002, UART-BR-003, UART-BR-005, UART-BR-006, UART-PAR-004, UART-PAR-005, UART-PAR-007, UART-REG-026, UART-REG-027',
     'PLANNED',''),
    ('COV-002','CTRL.PARITY',
     'All four parity modes (none/odd/even/mark) exercised independently in LOOP_EN loopback with parity error injection for odd and even modes.',
     STIM_STRAT['COV-002'],
     "2'b00 (none), 2'b01 (odd+error inject), 2'b10 (even+error inject), 2'b11 (mark)",
     'UART-FF-005, UART-FF-006, UART-FF-007, UART-FF-008, UART-FF-009, UART-REG-009, UART-VER-009',
     'PLANNED',''),
    ('COV-003','CTRL.STOP_BITS',
     'Both stop bit configurations (1 and 2 stop bits) exercised in loopback; receive accepts 1 stop bit in both modes.',
     STIM_STRAT['COV-003'],
     "1'b0 (1 stop bit), 1'b1 (2 stop bits)",
     'UART-FF-002, UART-FF-003, UART-FF-004, UART-REG-010',
     'PLANNED',''),
    ('COV-004','FIFO_STATUS.RX_LEVEL',
     'RX FIFO occupancy swept to key boundary values relative to RX_THRESH. G_FIFO_DEPTH used for instantiation independence.',
     STIM_STRAT['COV-004'],
     '0 (RX_EMPTY), RX_THRESH-1, RX_THRESH, RX_THRESH+1 (interrupt fires), G_FIFO_DEPTH (full/overrun)',
     'UART-FIFO-004, UART-FIFO-005, UART-FIFO-007, UART-FIFO-009, UART-REG-019, UART-REG-034',
     'PLANNED',''),
    ('COV-005','FIFO_STATUS.TX_LEVEL',
     'TX FIFO occupancy swept to key boundary values. TX_THRESH fires when occupancy falls BELOW threshold (asymmetric vs RX — common RTL bug source).',
     STIM_STRAT['COV-005'],
     '0 (TX_EMPTY), TX_THRESH-1, TX_THRESH, TX_THRESH+1 (interrupt fires), G_FIFO_DEPTH (full)',
     'UART-FIFO-002, UART-FIFO-003, UART-FIFO-006, UART-FIFO-008, UART-REG-016, UART-REG-017, UART-REG-033',
     'PLANNED',''),
    ('COV-006','TIMEOUT_VAL',
     'Timeout counter exercised at disable boundary (0) and active values including reset default. G_TIMEOUT_DEFAULT used for instantiation independence.',
     STIM_STRAT['COV-006'],
     '0x0000 (disabled), 0x0001 (minimum active), G_TIMEOUT_DEFAULT (reset), 0xFFFE, 0xFFFF (maximum)',
     'UART-TO-001, UART-TO-002, UART-TO-003, UART-TO-004, UART-TO-005, UART-TO-006, UART-REG-035, UART-REG-036, UART-PAR-006, UART-VER-008',
     'PLANNED',''),
    ('COV-007','INT_STATUS each_bit_independently',
     'Each of 8 interrupt sources set and cleared independently. Isolation before co-occurrence per NOTE-004.',
     STIM_STRAT['COV-007'],
     'Each of 8 bits: TIMEOUT, TX_THRESH, RX_THRESH, TX_EMPTY, RX_FULL, PARITY_ERR, FRAME_ERR, OVERRUN — set and W1C cleared independently',
     'UART-INT-001, UART-INT-002, UART-INT-003, UART-INT-004, UART-INT-005, UART-INT-006, UART-INT-007, UART-INT-008, UART-REG-037, UART-REG-039, UART-REG-040, UART-VER-004',
     'PLANNED',''),
    ('COV-008','INT_STATUS.OVERRUN && INT_STATUS.RX_FULL',
     'RX overrun while FIFO full — critical co-occurrence likely to expose FIFO pointer corruption bugs. Verifies overrun byte discarded and FIFO content preserved.',
     STIM_STRAT['COV-008'],
     'RX FIFO at exactly G_FIFO_DEPTH, then one additional received byte',
     'UART-FIFO-007, UART-INT-005, UART-INT-006, UART-REG-018, UART-REG-024, UART-VER-005, UART-VER-010',
     'PLANNED',''),
    ('COV-009','INT_STATUS.TIMEOUT && INT_STATUS.RX_THRESH',
     'Timeout pending simultaneously with RX threshold. Verifies IRQ OR reduction logic handles multiple sources correctly.',
     STIM_STRAT['COV-009'],
     'RX FIFO > RX_THRESH (threshold set), then idle until timeout fires while threshold remains set',
     'UART-TO-005, UART-INT-002, UART-INT-008, UART-REG-014',
     'PLANNED',''),
    ('COV-010','AXI_BRESP',
     "Both OKAY (2'b00) and SLVERR (2'b10) response codes on write channel. SLVERR via undefined offset 0x30 per NOTE-005.",
     STIM_STRAT['COV-010'],
     "Valid offset BRESP=OKAY, undefined offset 0x30 BRESP=SLVERR",
     'UART-IF-006, UART-IF-008, UART-REG-001',
     'PLANNED',''),
    ('COV-011','AXI_RRESP',
     "Both OKAY (2'b00) and SLVERR (2'b10) response codes on read channel. SLVERR via undefined offset 0x30 per NOTE-005.",
     STIM_STRAT['COV-011'],
     "Valid offset RRESP=OKAY, undefined offset 0x30 RRESP=SLVERR",
     'UART-IF-007, UART-IF-009, UART-REG-001',
     'PLANNED',''),
    ('COV-012','AXI_WRITE_ORDER',
     'Both AWVALID-first and WVALID-first AXI write orderings exercised. Many AXI-Lite slave implementations incorrectly assume AWVALID first.',
     STIM_STRAT['COV-012'],
     'AWVALID_FIRST ordering, WVALID_FIRST ordering',
     'UART-IF-005, UART-VER-006',
     'PLANNED',''),
    ('COV-013','BAUD_TUNING_WRITE_WHILE_ENABLED',
     'Write to BAUD_TUNING while UART_EN=1. Verifies silent-ignore behavior. Gap demo coverage point — closes when UART-BR-004 added to .req [confirmed].',
     STIM_STRAT['COV-013'],
     'UART_EN=1, write new value, readback must equal original value',
     'UART-REG-028, UART-VER-007',
     'PLANNED','Gap demo item — intentionally absent from initial .req cross-references'),
    ('COV-014','SCRATCH',
     'Walking-ones pattern plus all-zeros and all-ones written and read back. First test in every sequence per NOTE-006.',
     STIM_STRAT['COV-014'],
     '0x00000001..0x80000000 (walking ones), 0x000000FF, 0xFFFFFFFF (all-ones), 0x00000000 (all-zeros)',
     'UART-REG-044, UART-REG-045, UART-REG-046, UART-VER-003',
     'PLANNED','First test in all sequences per NOTE-006'),
    ('COV-015','RESET_VALUES',
     'All 12 registers read after reset deassertion and compared to spec-defined reset values from Table 8-x.',
     STIM_STRAT['COV-015'],
     '12 register reset values per spec Table 8-x',
     'UART-RST-001, UART-RST-002, UART-RST-003, UART-RST-004, UART-RST-005, UART-RST-006, UART-REG-004, UART-REG-013, UART-REG-031, UART-REG-036, UART-REG-046, UART-PAR-007, UART-PAR-008, UART-VER-002',
     'PLANNED',''),
    ('COV-016','UART_EN_CONTROL',
     'UART_EN, TX_EN, RX_EN, LOOP_EN control bit combinations verifying independent TX/RX path gating and loopback mode.',
     STIM_STRAT['COV-016'],
     'UART_EN=0 (halt), TX_EN=0 (TX off/RX active), RX_EN=0 (RX off/TX active), LOOP_EN=1 with UART_EN=1',
     'UART-EN-001, UART-EN-002, UART-EN-003, UART-EN-004, UART-EN-005, UART-EN-006, UART-REG-005, UART-REG-006, UART-REG-007, UART-REG-008',
     'PLANNED',''),
    ('COV-017','FRAME_STRUCTURAL',
     'Structural verification: 8-bit data frames, G_FIFO_DEPTH FIFO depth, 32-bit AXI data width, single-beat AXI transactions only.',
     STIM_STRAT['COV-017'],
     '8-bit frames; FIFO depth=G_FIFO_DEPTH; AXI data width=32; no burst transactions',
     'UART-FF-001, UART-FIFO-001, UART-IF-001, UART-IF-002, UART-IF-003, UART-IF-004, UART-IF-012, UART-IF-013, UART-IF-014, UART-PAR-001, UART-PAR-002, UART-PAR-009, UART-REG-032, UART-REG-033, UART-REG-034',
     'PLANNED',''),
    ('COV-018','READONLY_IGNORE',
     'Write to each read-only register; read back and verify state unchanged. Covers STATUS, FIFO_STATUS, INT_STATUS, INT_CLEAR, TX_DATA reads, RX_DATA writes.',
     STIM_STRAT['COV-018'],
     'Write all-ones to each RO register; compare pre- and post-write readback for each',
     'UART-IF-010, UART-IF-011, UART-REG-002, UART-REG-003, UART-REG-011, UART-REG-012, UART-REG-025, UART-REG-038, UART-REG-042, UART-REG-049, UART-REG-052',
     'PLANNED',''),
    ('COV-019','INT_THRESH_FRAME',
     'TX/RX threshold interrupts, FRAME_ERR, and TIMEOUT interrupt enable gate exercised via directed stimulus targeting INT_ENABLE masking and INT_STATUS bit positions.',
     STIM_STRAT['COV-019'],
     'TX FIFO below TX_THRESH; RX FIFO above RX_THRESH; stop-bit corruption for FRAME_ERR; TIMEOUT with INT_ENABLE gates',
     'UART-INT-009, UART-INT-010, UART-INT-011, UART-INT-012, UART-INT-013, UART-TO-007, UART-REG-015, UART-REG-020, UART-REG-021, UART-REG-022, UART-REG-023, UART-REG-029, UART-REG-030, UART-REG-037, UART-REG-039, UART-REG-040, UART-REG-041, UART-REG-043, UART-REG-047, UART-REG-048, UART-REG-050, UART-REG-051',
     'PLANNED',''),
]

for i, row_data in enumerate(cov_data, 3):
    for j, val in enumerate(row_data, 1):
        cell = ws_cg.cell(row=i, column=j, value=val)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

print(f"Coverage_Goals: wrote {len(cov_data)} rows")

# ── 6. Strategy_Notes tab ─────────────────────────────────────────────────────
ws_sn = wb['Strategy_Notes']
for row in ws_sn.iter_rows(min_row=3, max_row=max(ws_sn.max_row, 12)):
    for cell in row:
        cell.value = None

strategy_data = [
    ('NOTE-001','Primary Stimulus Method',
     'Loopback mode (LOOP_EN) is the primary stimulus strategy for all frame format and parity coverage goals. External UART stimulus is not used in the initial verification pass.',
     'COV-002, COV-003, COV-004, COV-005'),
    ('NOTE-002','NCO Verification',
     'The NCO baud rate generator is verified indirectly via loopback frame integrity at each COV-001 tuning word. Direct accumulator state observation is not required by any requirement.',
     'COV-001'),
    ('NOTE-003','FIFO Threshold Expressions',
     'FIFO threshold coverage goals COV-004 and COV-005 use relative expressions (RX_THRESH +/- 1) rather than absolute values so they remain valid across all G_FIFO_DEPTH instantiations without test modification.',
     'COV-004, COV-005'),
    ('NOTE-004','Interrupt Verification Order',
     'Interrupt model verification prioritizes isolation first (COV-007 each_bit_independently) then co-occurrence (COV-008, COV-009). A combined all-sources-simultaneously scenario is explicitly not targeted — it adds complexity without additional requirement closure.',
     'COV-007, COV-008, COV-009'),
    ('NOTE-005','AXI SLVERR Address',
     'AXI-Lite SLVERR coverage (COV-010, COV-011) is achieved by accessing offset 0x30 which is the first undefined address above the register map.',
     'COV-010, COV-011'),
    ('NOTE-006','SCRATCH Test Ordering',
     'The SCRATCH register walking-ones test (COV-014) is always the first test sequence executed before any functional register access, establishing AXI-Lite connectivity as a precondition for all subsequent tests.',
     'COV-014'),
    ('NOTE-007','Gap Demo Requirement',
     'UART-BR-004 is intentionally absent from the [confirmed] section of the .req file in the initial demonstration. This produces a Direction A ERROR in the gap report. Closing the gap by adding the [CONFIRMED] disposition to the .req file is the live demonstration of pssgen traceability. Do not add UART-BR-004 here until after the demo.',
     'UART-BR-004, COV-013'),
]

for i, row_data in enumerate(strategy_data, 3):
    for j, val in enumerate(row_data, 1):
        cell = ws_sn.cell(row=i, column=j, value=val)
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

print(f"Strategy_Notes: wrote {len(strategy_data)} rows")

# ── 7. Summary tab ────────────────────────────────────────────────────────────
ws_sum = wb['Summary']
for row in ws_sum.iter_rows(min_row=3, max_row=max(ws_sum.max_row, 20)):
    for cell in row:
        cell.value = None

families = ['PAR','IF','EN','BR','FF','FIFO','TO','INT','REG','RST','VER']
hdr_font  = Font(name='Arial', size=9, bold=True)

# Column headers
for j, h in enumerate(['Family','Total','Covered','Waived','Gap','% Complete'], 1):
    ws_sum.cell(row=2, column=j, value=h).font = hdr_font

for i, fam in enumerate(families):
    row = 3 + i
    total_f   = f'=COUNTIF(VPR!B4:B144,"{fam}")'
    covered_f = f'=COUNTIFS(VPR!B4:B144,"{fam}",VPR!X4:X144,"<>OPEN",VPR!X4:X144,"<>WAIVED")'
    waived_f  = f'=COUNTIFS(VPR!B4:B144,"{fam}",VPR!K4:K144,"WAIVED")'
    gap_f     = f'=B{row}-C{row}-D{row}'
    pct_f     = f'=IFERROR((C{row}+D{row})/B{row},0)'
    ws_sum.cell(row=row, column=1, value=fam).font = data_font
    ws_sum.cell(row=row, column=2, value=total_f).font = data_font
    ws_sum.cell(row=row, column=3, value=covered_f).font = data_font
    ws_sum.cell(row=row, column=4, value=waived_f).font = data_font
    ws_sum.cell(row=row, column=5, value=gap_f).font = data_font
    pct = ws_sum.cell(row=row, column=6, value=pct_f)
    pct.font = data_font
    pct.number_format = '0%'

total_row = 3 + len(families)
ws_sum.cell(row=total_row, column=1, value='TOTAL').font = hdr_font
ws_sum.cell(row=total_row, column=2, value=f'=SUM(B3:B{total_row-1})').font = hdr_font
ws_sum.cell(row=total_row, column=3, value=f'=SUM(C3:C{total_row-1})').font = hdr_font
ws_sum.cell(row=total_row, column=4, value=f'=SUM(D3:D{total_row-1})').font = hdr_font
ws_sum.cell(row=total_row, column=5, value=f'=SUM(E3:E{total_row-1})').font = hdr_font
pct_tot = ws_sum.cell(row=total_row, column=6, value=f'=IFERROR((C{total_row}+D{total_row})/B{total_row},0)')
pct_tot.font = hdr_font
pct_tot.number_format = '0%'

print(f"Summary: wrote {len(families)} family rows + TOTAL row")

# ── 8. Save ───────────────────────────────────────────────────────────────────
wb.save(DST)
print(f"\nSaved: {DST}")

# ── 9. Verification checks ────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(DST, data_only=False)
ws_v = wb2['VPR']

# Total rows including header(row2) + example(row3) + 141 data = 143 non-empty in col A
all_rows = sum(1 for row in ws_v.iter_rows(min_row=2, max_row=200, min_col=1, max_col=1)
               if row[0].value)
# Count data rows 4..144
data_rows = sum(1 for row in ws_v.iter_rows(min_row=4, max_row=144, min_col=1, max_col=1)
                if row[0].value)
# Count WAIVED (col K=11, data rows 4..144)
waived = sum(1 for row in ws_v.iter_rows(min_row=4, max_row=144, min_col=11, max_col=11)
             if row[0].value == 'WAIVED')
# Count open-not-waived (col H blank, col K not WAIVED)
open_not_waived = []
for row in ws_v.iter_rows(min_row=4, max_row=144):
    if not row[7].value and row[10].value != 'WAIVED':
        open_not_waived.append(row[0].value)
# Count formula rows (col X=24, rows 3..144)
formula_count = sum(1 for row in ws_v.iter_rows(min_row=3, max_row=144, min_col=24, max_col=24)
                    if row[0].value and str(row[0].value).startswith('='))

print(f"\n--- Verification Checks ---")
print(f"  VPR rows with Req_ID (rows 2..200): {all_rows}  (expected 143 = 1 header + 1 example + 141 data)")
print(f"  Data rows (rows 4..144): {data_rows}  (expected 141)")
print(f"  WAIVED rows: {waived}  (expected 2)")
print(f"  Open (no Covered_By, not WAIVED): {len(open_not_waived)}  (expected 1 = UART-BR-004)")
print(f"  Overall_Status formulas (rows 3..144): {formula_count}  (expected 142)")
if open_not_waived:
    print(f"  Open rows: {open_not_waived}")
ok = (all_rows == 143 and data_rows == 141 and waived == 2 and
      len(open_not_waived) == 1 and open_not_waived[0] == 'UART-BR-004' and
      formula_count == 142)
print(f"\n  {'ALL CHECKS PASS' if ok else 'CHECKS FAILED'}")
