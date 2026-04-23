# ===========================================================
# FILE:         agents/results_collector.py
# PROJECT:      pssgen — AI-Driven PSS + UVM + C Testbench Generator
# COPYRIGHT:    Copyright (c) 2026 BelTech Systems LLC
# LICENSE:      MIT License — see LICENSE file for details
# ===========================================================
#
# DESCRIPTION:
#   Reads xsim simulation results, writes RTL execution results back to
#   the VPR spreadsheet (RTL_Status, RTL_Run_Date, RTL_Commit,
#   RTL_Evidence), and generates a gap_report.json for Grafana dashboards.
#   Implements the --collect-results pipeline mode (OI-29).
#
# LAYER:        3 — agents
# PHASE:        v1a
#
# FUNCTIONS:
#   parse_xsim_log(log_path)
#     Parse an xsim.log file and return a SimResult dataclass.
#   parse_coverage_data(sim_log_path)
#     Parse xsim.log for detailed CAE coverage data; return dict with
#     sequences, SLVERR events, scoreboard, and per-sequence timeouts.
#   write_vpr_results(vplan_path, sim_result, req_ids)
#     Write RTL_* columns back to VPR Tab 1; return count of rows updated.
#   generate_gap_report_json(vplan_path, sim_result, out_path)
#     Read VPR and generate gap_report.json; return the path written.
#
# DEPENDENCIES:
#   Standard library:  dataclasses, datetime, json, os, re, subprocess
#   Internal:          parser.vplan_parser
#   Third-party:       openpyxl
#
# HISTORY:
#   v1a  2026-04-16  SB  Initial implementation; --collect-results pipeline (OI-29)
#   v1b  2026-04-17  SB  Add family_summary block to gap_report.json (Grafana bar chart)
#   v1c  2026-04-23  SB  Add parse_coverage_data() — CAE-002 sequence/SLVERR parser
#
# ===========================================================
"""agents/results_collector.py — Simulation results collection and VPR write-back.

Phase: v1a
Layer: 3 (agents)

Parses xsim.log files, writes RTL execution results back to the VPR
spreadsheet, and generates gap_report.json for Grafana dashboards.
"""
from __future__ import annotations

import json
import os
import re
import subprocess
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from typing import Optional

import openpyxl

from parser.vplan_parser import parse_vplan


# ── Skip identifiers — non-data rows in the VPR tab ─────────────────────────
_VPR_SKIP_IDS: frozenset[str] = frozenset({"Req_ID", "[BLOCK-FAM-NNN]", ""})


@dataclass
class SimResult:
    """Result of parsing a single xsim simulation log.

    Attributes:
        log_path: Absolute or relative path to the parsed log file.
        passed: True when UVM_ERROR == 0 and UVM_FATAL == 0.
        uvm_errors: Count of UVM_ERROR messages from the summary block.
        uvm_fatals: Count of UVM_FATAL messages from the summary block.
        uvm_warnings: Count of UVM_WARNING messages from the summary block.
        uvm_infos: Count of UVM_INFO messages from the summary block.
        sim_time_ns: Final simulation time in nanoseconds; 0.0 if unknown.
        commit_hash: Short git commit hash at parse time; "unknown" if git unavailable.
        run_date: ISO-8601 date string of the run (today when parse_xsim_log is called).
        coverage_pct: Overall functional coverage percentage; 0.0 if not reported.
        log_lines: Last 50 lines of the log file for evidence archiving.
    """
    log_path: str
    passed: bool
    uvm_errors: int
    uvm_fatals: int
    uvm_warnings: int
    uvm_infos: int
    sim_time_ns: float
    commit_hash: str
    run_date: str
    coverage_pct: float
    log_lines: list[str] = field(default_factory=list)


# ── Regex patterns for xsim.log parsing ──────────────────────────────────────
# Matches summary-block lines like "UVM_ERROR :    0" (variable whitespace).
_RE_UVM_SEVERITY = re.compile(
    r"^UVM_(INFO|WARNING|ERROR|FATAL)\s*:\s+(\d+)\s*$"
)
# Matches "$finish called at time : 1525 ns" or "1525.000 ns"
_RE_SIM_TIME = re.compile(
    r"\$finish called at time\s*:\s*([\d.]+)\s*ns", re.IGNORECASE
)
# Matches "[COV] ... coverage: 41.7%"
_RE_COVERAGE = re.compile(r"\[COV\].*?coverage:\s*([\d.]+)%", re.IGNORECASE)


def _get_commit_hash() -> str:
    """Return the short git commit hash of HEAD, or 'unknown' if git unavailable.

    Returns:
        7-character short hash string, or 'unknown'.
    """
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True,
            text=True,
            timeout=5,
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except (FileNotFoundError, subprocess.TimeoutExpired, OSError):
        pass
    return "unknown"


def parse_xsim_log(log_path: str) -> SimResult:
    """Parse an xsim.log file and return a SimResult.

    Extracts UVM severity counts from the UVM Report Summary block,
    simulation end time from the $finish line, and coverage percentage
    from [COV] report lines. Runs git rev-parse to capture the current
    commit hash.

    Args:
        log_path: Path to the xsim.log file to parse.

    Returns:
        SimResult with all fields populated. passed is True only when
        both uvm_errors and uvm_fatals are zero.
    """
    uvm_counts: dict[str, int] = {
        "INFO": 0,
        "WARNING": 0,
        "ERROR": 0,
        "FATAL": 0,
    }
    sim_time_ns: float = 0.0
    coverage_pct: float = 0.0

    with open(log_path, "r", encoding="utf-8", errors="replace") as fh:
        all_lines = fh.readlines()

    for line in all_lines:
        stripped = line.rstrip()

        # UVM summary block severities (summary lines have no path prefix)
        m = _RE_UVM_SEVERITY.match(stripped)
        if m:
            uvm_counts[m.group(1)] = int(m.group(2))
            continue

        # Simulation end time
        m = _RE_SIM_TIME.search(stripped)
        if m:
            try:
                sim_time_ns = float(m.group(1))
            except ValueError:
                pass

        # Coverage percentage from [COV] subscriber report line
        m = _RE_COVERAGE.search(stripped)
        if m:
            try:
                coverage_pct = float(m.group(1))
            except ValueError:
                pass

    log_lines = [ln.rstrip() for ln in all_lines[-50:]]

    return SimResult(
        log_path=log_path,
        passed=(uvm_counts["ERROR"] == 0 and uvm_counts["FATAL"] == 0),
        uvm_errors=uvm_counts["ERROR"],
        uvm_fatals=uvm_counts["FATAL"],
        uvm_warnings=uvm_counts["WARNING"],
        uvm_infos=uvm_counts["INFO"],
        sim_time_ns=sim_time_ns,
        commit_hash=_get_commit_hash(),
        run_date=date.today().isoformat(),
        coverage_pct=coverage_pct,
        log_lines=log_lines,
    )


def write_vpr_results(
    vplan_path: str,
    sim_result: SimResult,
    req_ids: Optional[list[str]] = None,
) -> int:
    """Write RTL execution results back to the VPR spreadsheet Tab 1.

    Locates the VPR tab, reads column names dynamically from header row 2,
    then writes RTL_Status / RTL_Run_Date / RTL_Commit / RTL_Evidence for
    every qualifying data row. WAIVED rows are always skipped. Rows that
    already carry RTL_Status = "PASS" are never downgraded.

    Args:
        vplan_path: Path to the VPR spreadsheet (.xlsx). Modified in place.
        sim_result: Parsed simulation result to write back.
        req_ids: If provided, only update rows whose Req_ID is in this list.
                 If None, update all non-WAIVED rows.

    Returns:
        Count of rows updated.

    Raises:
        ValueError: If the VPR tab is missing or required RTL_* columns are absent.
    """
    # Must open with data_only=False so formula cells are preserved on save.
    wb = openpyxl.load_workbook(vplan_path, read_only=False, data_only=False)

    if "VPR" not in wb.sheetnames:
        raise ValueError(f"VPR tab not found in {vplan_path}")

    ws = wb["VPR"]

    # ── Discover column positions from header row 2 ───────────────────────────
    col_map: dict[str, int] = {}  # header name → 0-based column index
    header_row = next(ws.iter_rows(min_row=2, max_row=2))
    for idx, cell in enumerate(header_row):
        if cell.value:
            col_map[str(cell.value).strip()] = idx

    required = {"RTL_Status", "RTL_Run_Date", "RTL_Commit", "RTL_Evidence"}
    missing = required - col_map.keys()
    if missing:
        raise ValueError(
            f"VPR tab missing required column header(s): {', '.join(sorted(missing))}"
        )

    req_id_col      = col_map.get("Req_ID", 0)
    disposition_col = col_map.get("Disposition", 10)
    rtl_status_col  = col_map["RTL_Status"]
    rtl_date_col    = col_map["RTL_Run_Date"]
    rtl_commit_col  = col_map["RTL_Commit"]
    rtl_evidence_col = col_map["RTL_Evidence"]

    new_status = "PASS" if sim_result.passed else "FAIL"
    req_ids_set = set(req_ids) if req_ids is not None else None
    rows_updated = 0

    for row in ws.iter_rows(min_row=4):
        # ── Req_ID guard ─────────────────────────────────────────────────────
        raw_id = row[req_id_col].value
        req_id = str(raw_id).strip() if raw_id is not None else ""
        if req_id in _VPR_SKIP_IDS:
            continue

        # ── Optional req_ids filter ──────────────────────────────────────────
        if req_ids_set is not None and req_id not in req_ids_set:
            continue

        # ── Skip WAIVED rows ─────────────────────────────────────────────────
        raw_disp = row[disposition_col].value if disposition_col < len(row) else None
        disposition = str(raw_disp).strip().upper() if raw_disp else ""
        if disposition == "WAIVED":
            continue

        # ── Never downgrade an existing PASS ────────────────────────────────
        if rtl_status_col < len(row):
            existing = row[rtl_status_col].value
            existing_str = str(existing).strip() if existing is not None else ""
            if existing_str == "PASS" and new_status != "PASS":
                continue

        # ── Write RTL_* columns ───────────────────────────────────────────────
        row[rtl_status_col].value   = new_status
        row[rtl_date_col].value     = sim_result.run_date
        row[rtl_commit_col].value   = sim_result.commit_hash
        row[rtl_evidence_col].value = sim_result.log_path
        rows_updated += 1

    wb.save(vplan_path)
    wb.close()

    print(f"[results_collector] VPR write-back: {rows_updated} row(s) updated -> {vplan_path}")
    return rows_updated


def generate_gap_report_json(
    vplan_path: str,
    sim_result: SimResult,
    out_path: str,
) -> str:
    """Read VPR and generate gap_report.json for Grafana.

    Builds a structured JSON report containing the simulation result,
    per-requirement status, and summary counts derived from RTL_Status
    values after write-back. Overall_Status is read from the formula
    result column (col 27) using data_only=True.

    Args:
        vplan_path: Path to the VPR spreadsheet (.xlsx).
        sim_result: Parsed simulation result.
        out_path: Destination path for gap_report.json.

    Returns:
        The out_path that was written.
    """
    # ── Parse requirements and coverage items (data_only for formula results) ─
    vplan_result = parse_vplan(vplan_path)

    # ── Read per-row Overall_Status from the spreadsheet ─────────────────────
    # parse_vplan opens with data_only=True — reuse that behaviour.
    wb_data = openpyxl.load_workbook(vplan_path, read_only=True, data_only=True)
    ws_data = wb_data["VPR"]

    # Build column map from header row 2
    col_map_data: dict[str, int] = {}
    all_rows = list(ws_data.iter_rows())
    if len(all_rows) >= 2:
        for idx, cell in enumerate(all_rows[1]):   # row 2 (0-based index 1)
            if cell.value:
                col_map_data[str(cell.value).strip()] = idx

    req_id_col_d      = col_map_data.get("Req_ID", 0)
    rtl_status_col_d  = col_map_data.get("RTL_Status")
    overall_col_d     = col_map_data.get("Overall_Status")
    _SKIP = frozenset({"Req_ID", "[BLOCK-FAM-NNN]", ""})

    status_map: dict[str, dict[str, str]] = {}
    for row in all_rows[3:]:   # data rows from row 4 (0-based index 3)
        if req_id_col_d >= len(row):
            continue
        raw = row[req_id_col_d].value
        req_id = str(raw).strip() if raw is not None else ""
        if req_id in _SKIP:
            continue

        rtl_val = ""
        if rtl_status_col_d is not None and rtl_status_col_d < len(row):
            v = row[rtl_status_col_d].value
            rtl_val = str(v).strip() if v is not None else ""

        overall_val = ""
        if overall_col_d is not None and overall_col_d < len(row):
            v = row[overall_col_d].value
            overall_val = str(v).strip() if v is not None else ""

        status_map[req_id] = {
            "rtl_status":     rtl_val,
            "overall_status": overall_val,
        }

    wb_data.close()

    # ── Build requirements list and summary counts ────────────────────────────
    requirements: list[dict] = []
    summary: dict[str, int] = {
        "total":    0,
        "waived":   0,
        "passing":  0,
        "failing":  0,
        "not_run":  0,
        "at_risk":  0,
        "open":     0,
    }

    for req_id, req_detail in vplan_result.requirements.items():
        is_waived = req_detail.get("waived", False)
        row_status = status_map.get(req_id, {})
        rtl_status = row_status.get("rtl_status", "NOT_RUN")
        overall_status = row_status.get("overall_status", "") or rtl_status

        disposition = "WAIVED" if is_waived else req_detail.get("disposition", "GENERATED")

        requirements.append({
            "req_id":         req_id,
            "family":         req_detail.get("family", ""),
            "disposition":    disposition,
            "covered_by":     req_detail.get("covered_by", ""),
            "rtl_status":     rtl_status,
            "overall_status": overall_status,
        })

        summary["total"] += 1
        if is_waived:
            summary["waived"] += 1
        elif rtl_status == "PASS":
            summary["passing"] += 1
        elif rtl_status == "FAIL":
            summary["failing"] += 1
        else:
            summary["not_run"] += 1

    # ── Build family_summary: per-family pass/fail/waived/not_run counts ────────
    family_buckets: dict[str, dict[str, int]] = {}
    for req in requirements:
        fam = req.get("family") or ""
        if fam not in family_buckets:
            family_buckets[fam] = {
                "total": 0, "passing": 0, "failing": 0,
                "waived": 0, "not_run": 0,
            }
        b = family_buckets[fam]
        b["total"] += 1
        if req["disposition"] == "WAIVED":
            b["waived"] += 1
        elif req["rtl_status"] == "PASS":
            b["passing"] += 1
        elif req["rtl_status"] == "FAIL":
            b["failing"] += 1
        else:
            b["not_run"] += 1
    family_summary = {fam: family_buckets[fam] for fam in sorted(family_buckets)}
    family_summary_array = [
        {"family": k, **v}
        for k, v in sorted(family_summary.items())
    ]

    report = {
        "generated": sim_result.run_date,
        "commit":    sim_result.commit_hash,
        "sim_result": {
            "passed":       sim_result.passed,
            "uvm_errors":   sim_result.uvm_errors,
            "uvm_fatals":   sim_result.uvm_fatals,
            "coverage_pct": sim_result.coverage_pct,
        },
        "summary":              summary,
        "family_summary":       family_summary,
        "family_summary_array": family_summary_array,
        "requirements":         requirements,
    }

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, indent=2)

    return out_path


# ── Regex patterns for parse_coverage_data ───────────────────────────────────

# Full UVM message line — optional file-path token before @.
# Handles both "UVM_INFO @ 0:" and "UVM_INFO path/file.sv(N) @ 0:".
_RE_UVM_MSG = re.compile(
    r"^(UVM_INFO|UVM_WARNING|UVM_ERROR|UVM_FATAL)\s+"
    r"(?:[^@\s]\S*\s+)?"       # optional file-path token (must not start with @)
    r"@\s*([\d.]+)\s*:\s+"     # @ <time_ps>:
    r"(\S+)\s+"                # component
    r"\[([^\]]+)\]\s*(.*)",    # [tag] message
    re.IGNORECASE,
)
# Sequence number from component path: @@cov<NNN>
_RE_SEQ_NUM = re.compile(r"@@cov(\d+)", re.IGNORECASE)
# Severity summary line from UVM Report Summary block
_RE_SEV_SUMMARY = re.compile(r"^UVM_(INFO|WARNING|ERROR|FATAL)\s*:\s+(\d+)\s*$")
# Poll timeout fields — handles both em-dash and hyphen separators
_RE_TIMEOUT = re.compile(
    r"axi_poll\s+timeout.*?addr=(0x[\da-fA-F]+)\s+"
    r"mask=(0x[\da-fA-F]+)\s+exp=(0x[\da-fA-F]+)\s+got=(0x[\da-fA-F]+)",
    re.IGNORECASE,
)
# SLVERR event: [SB] read|write SLVERR — addr=0x... resp=...
_RE_SLVERR = re.compile(
    r"\[SB\]\s+(read|write)\s+SLVERR\s*(?:—|-+)\s*"
    r"addr=(0x[\da-fA-F]+)\s+resp=(\S+)",
    re.IGNORECASE,
)
# Scoreboard check_phase error count
_RE_SB_ERRORS = re.compile(r"Scoreboard check_phase:\s*(\d+)\s*error", re.IGNORECASE)
# Shadow register map disabled warning — matches truncated "availabl..." in BALU log
_RE_SHADOW = re.compile(r"Shadow register map not avail", re.IGNORECASE)


def parse_coverage_data(sim_log_path: str) -> dict:
    """Parse an xsim.log file and return detailed CAE coverage data.

    Extracts UVM severity counts, coverage percentage, per-sequence status
    (PASS/TIMEOUT/ERROR/NOT_RUN), poll-timeout details, SLVERR events,
    and scoreboard state. Always returns entries for all 19 sequences
    (COV-001 through COV-019), using NOT_RUN for those with no log activity.

    UVM message timestamps are in picoseconds; completion_time_ns divides
    by 1000. simulation_time_ns is read directly from the $finish line (ns).

    Args:
        sim_log_path: Path to the xsim.log file to parse.

    Returns:
        Dict conforming to the CAE-002 schema.
    """
    # Per-sequence accumulator: keys 1–19
    seq_state: dict[int, dict] = {
        n: {
            "tag": None,
            "messages": [],
            "timeouts": [],
            "has_error": False,
            "last_time_ps": 0.0,
        }
        for n in range(1, 20)
    }

    uvm_counts: dict[str, int] = {"INFO": 0, "WARNING": 0, "ERROR": 0, "FATAL": 0}
    simulation_time_ns: float = 0.0
    coverage_pct: float = -1.0
    slverr_events: list[dict] = []
    sb_errors: int = 0
    sb_disabled: bool = False
    assertions_fired: list[dict] = []
    in_summary: bool = False

    with open(sim_log_path, "r", encoding="utf-8", errors="replace") as fh:
        for raw_line in fh:
            line = raw_line.rstrip()

            # ── UVM Report Summary block ──────────────────────────────────
            if "--- UVM Report Summary ---" in line:
                in_summary = True

            if in_summary:
                sm = _RE_SEV_SUMMARY.match(line)
                if sm:
                    uvm_counts[sm.group(1)] = int(sm.group(2))

            # ── Simulation end time ($finish line) ────────────────────────
            m = _RE_SIM_TIME.search(line)
            if m:
                simulation_time_ns = float(m.group(1))

            # ── Coverage percentage ───────────────────────────────────────
            m = _RE_COVERAGE.search(line)
            if m:
                coverage_pct = float(m.group(1))

            # ── Shadow register map disabled ──────────────────────────────
            if _RE_SHADOW.search(line):
                sb_disabled = True

            # ── Parse UVM message lines ───────────────────────────────────
            m = _RE_UVM_MSG.match(line)
            if not m:
                continue

            sev = m.group(1)          # e.g. "UVM_INFO"
            time_ps = float(m.group(2)) if m.group(2) else 0.0
            component = m.group(3)
            tag = m.group(4)
            msg = m.group(5)

            # ── Scoreboard check_phase ────────────────────────────────────
            sbm = _RE_SB_ERRORS.search(msg)
            if sbm:
                sb_errors = int(sbm.group(1))

            # ── SLVERR event ──────────────────────────────────────────────
            slv = _RE_SLVERR.search(line)
            if slv:
                slverr_events.append({
                    "direction": slv.group(1).lower(),
                    "addr": slv.group(2),
                    "resp": slv.group(3),
                    "time_ns": time_ps / 1000.0,
                })

            # ── Sequence data (component must contain @@covNNN) ───────────
            seq_m = _RE_SEQ_NUM.search(component)
            if not seq_m:
                continue
            n = int(seq_m.group(1))
            if not (1 <= n <= 19):
                continue

            ss = seq_state[n]
            if ss["tag"] is None:
                ss["tag"] = tag
            ss["last_time_ps"] = max(ss["last_time_ps"], time_ps)

            if sev == "UVM_INFO":
                ss["messages"].append(msg)
            elif sev in ("UVM_ERROR", "UVM_FATAL"):
                ss["has_error"] = True

            # Timeout in this line?
            tm = _RE_TIMEOUT.search(line)
            if tm:
                ss["timeouts"].append({
                    "addr": tm.group(1),
                    "mask": tm.group(2),
                    "expected": tm.group(3),
                    "got": tm.group(4),
                })

    # ── Build ordered sequences list (COV-001 through COV-019) ───────────────
    sequences: list[dict] = []
    for n in range(1, 20):
        ss = seq_state[n]
        tag = ss["tag"] or f"RCOV{n:03d}"

        if ss["has_error"]:
            status = "ERROR"
        elif ss["timeouts"]:
            status = "TIMEOUT"
        elif ss["messages"]:
            status = "PASS"
        else:
            status = "NOT_RUN"

        sequences.append({
            "seq_id": f"COV-{n:03d}",
            "tag": tag,
            "status": status,
            "messages": ss["messages"],
            "timeouts": ss["timeouts"],
            "completion_time_ns": ss["last_time_ps"] / 1000.0,
        })

    return {
        "sim_log": sim_log_path,
        "parsed_at": datetime.now(timezone.utc).isoformat(),
        "simulation_time_ns": simulation_time_ns,
        "uvm_counts": {
            "info": uvm_counts["INFO"],
            "warning": uvm_counts["WARNING"],
            "error": uvm_counts["ERROR"],
            "fatal": uvm_counts["FATAL"],
        },
        "coverage_pct": coverage_pct,
        "sequences": sequences,
        "slverr_events": slverr_events,
        "scoreboard": {
            "errors": sb_errors,
            "disabled": sb_disabled,
        },
        "assertions_fired": assertions_fired,
    }
