# ===========================================================
# FILE:         parser/regmap_parser.py
# PROJECT:      pssgen — AI-Driven PSS + UVM + C Testbench Generator
# COPYRIGHT:    Copyright (c) 2026 BelTech Systems LLC
# LICENSE:      MIT License — see LICENSE file for details
# ===========================================================
#
# DESCRIPTION:
#   Reads register map data from either an .xlsx spreadsheet (the four-sheet
#   pssgen register map format, single-sheet simple_block format, or multi-file
#   system format) or a plain English .intent file containing a "register map:"
#   section. Returns a normalized register_map dict stored in ir.register_map.
#
# LAYER:        1 — parser
# PHASE:        v4c
#
# FUNCTIONS:
#   parse_regmap(source_file)
#     Dispatch to correct parser based on extension and detected format.
#     Returns normalized register_map dict.
#   detect_regmap_format(xlsx_path)
#     Auto-detect format: "simple_block", "full_block", or "system".
#   _parse_xlsx(source_file)
#     Read Globals, Blocks, RegisterMap, and Enums sheets from full_block xlsx.
#   _parse_simple_block(xlsx_path)
#     Parse single-sheet 15-18 column simple_block spreadsheet.
#   _parse_system(xlsx_path)
#     Parse system spreadsheet referencing multiple block files.
#   _parse_intent_regmap(content)
#     Extract a register map: section from plain English intent content.
#
# DEPENDENCIES:
#   Standard library:  re, os
#   Internal:          parser.dispatch (ParseError)
#
# HISTORY:
#   v4a   2026-04-03  SB  Initial implementation; xlsx + plain English register map parsing
#   v4c   2026-04-05  SB  Added detect_regmap_format, _parse_simple_block, _parse_system;
#                         updated parse_regmap dispatch for three formats
#
# ===========================================================
"""parser/regmap_parser.py — Register map spreadsheet and intent section parser.

Phase: v4c
Layer: 1 (parser)

Parses register map data from a pssgen-format .xlsx spreadsheet (full_block,
simple_block, or system format) or from a ``register map:`` section embedded
in a .intent file. The returned dict has four keys (globals, blocks, registers,
enums) and is stored in ir.register_map.
"""
import re
import os

from parser.verilog import ParseError


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_regmap(source_file: str) -> dict:
    """Parse a register map from spreadsheet or intent.

    Dispatches to the appropriate parser based on extension and detected
    spreadsheet format. Supports full_block (four-sheet), simple_block
    (single-sheet 15–18 column), system (multi-file reference), and
    .intent plain English formats.

    Args:
        source_file: Path to ``.xlsx`` or ``.intent`` file.

    Returns:
        register_map dict with keys: ``globals``, ``blocks``,
        ``registers``, ``enums``.

    Raises:
        ValueError: If the file extension is not ``.xlsx`` or ``.intent``.
        ParseError: If the file cannot be parsed.
    """
    ext = os.path.splitext(source_file)[1].lower()
    if ext == ".intent":
        with open(source_file, "r", encoding="utf-8") as fh:
            content = fh.read()
        return _parse_intent_regmap(content)
    elif ext == ".xlsx":
        fmt = detect_regmap_format(source_file)
        if fmt == "simple_block":
            return _parse_simple_block(source_file)
        if fmt == "system":
            return _parse_system(source_file)
        return _parse_xlsx(source_file)  # full_block
    else:
        raise ValueError(
            f"register map file must be .xlsx or .intent, got '{ext}'"
        )


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

def detect_regmap_format(xlsx_path: str) -> str:
    """Detect register map spreadsheet format.

    Inspects sheet names (and for ambiguous cases the header row of the Blocks
    sheet) to classify the workbook as one of three known formats.

    Detection rules applied in order:
      1. Has ``RegisterMap`` sheet → ``"full_block"``
      2. Has ``Blocks`` sheet:
           If Blocks header contains ``spreadsheet_file`` column → ``"system"``
           Else → ``"full_block"``
      3. Single sheet — check header row for ``spreadsheet_file`` → ``"system"``
      4. Fallback → ``"simple_block"``

    Args:
        xlsx_path: Path to the ``.xlsx`` file to inspect.

    Returns:
        Format string: ``"simple_block"``, ``"full_block"``, or ``"system"``.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ParseError(
            "openpyxl is required. Install with: pip install openpyxl"
        ) from exc

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"Cannot open workbook: {exc}") from exc

    sheet_names = wb.sheetnames
    wb.close()

    # Rule 1: has a RegisterMap sheet AND at least one companion sheet → full_block.
    # A single-sheet workbook named "RegisterMap" is treated as simple_block (falls
    # through to rule 4) so that the 18-column simple format can share the sheet name.
    if "RegisterMap" in sheet_names and len(sheet_names) > 1:
        return "full_block"

    # Rule 2: has a Blocks sheet
    if "Blocks" in sheet_names:
        try:
            wb2 = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb2["Blocks"]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True),
                          None)
            wb2.close()
        except Exception:
            return "full_block"
        if header:
            header_lower = [str(c).lower().strip() if c else "" for c in header]
            if "spreadsheet_file" in header_lower:
                return "system"
        return "full_block"

    # Rule 3: single sheet — check its header for spreadsheet_file
    if len(sheet_names) == 1:
        try:
            wb3 = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb3[sheet_names[0]]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True),
                          None)
            wb3.close()
        except Exception:
            return "simple_block"
        if header:
            header_lower = [str(c).lower().strip() if c else "" for c in header]
            if "spreadsheet_file" in header_lower:
                return "system"

    # Rule 4: fallback
    return "simple_block"


# ---------------------------------------------------------------------------
# XLSX parser
# ---------------------------------------------------------------------------

def _parse_xlsx(source_file: str) -> dict:
    """Read a pssgen-format register map workbook.

    Requires the ``openpyxl`` package. Opens the workbook in read-only mode
    to minimise memory use on large spreadsheets.

    Args:
        source_file: Path to the ``.xlsx`` workbook.

    Returns:
        Normalised register_map dict.

    Raises:
        ParseError: If required sheets are missing or openpyxl is unavailable.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ParseError(
            "openpyxl is required for register map parsing. "
            "Install it with: pip install openpyxl"
        ) from exc

    try:
        wb = load_workbook(source_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"Cannot open register map workbook: {exc}") from exc

    globals_dict = _read_globals_sheet(wb)
    blocks_list = _read_blocks_sheet(wb)
    registers_list = _read_registermap_sheet(wb)
    enums_dict = _read_enums_sheet(wb)

    wb.close()

    return {
        "globals": globals_dict,
        "blocks": blocks_list,
        "registers": registers_list,
        "enums": enums_dict,
    }


def _is_example_row(value: str | None) -> bool:
    """Return True if the cell value marks a template example row."""
    if value is None:
        return False
    return str(value).startswith("[EXAMPLE")


def _str_or_none(value) -> str | None:
    """Convert a cell value to a stripped string, or None if blank."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _read_globals_sheet(wb) -> dict:
    """Extract the Globals sheet into a flat key→value dict."""
    if "Globals" not in wb.sheetnames:
        return {}

    ws = wb["Globals"]
    result: dict = {}
    # Skip row 1 (cover note or header). Read from row 2 onwards.
    # For workbooks without the cover note, row 2 is still the header — we
    # detect the header by checking whether the first cell says "Key".
    # Strategy: read all rows, skip rows where key is None or "Key" (header)
    # or starts with "pssgen Register" (cover note), or starts with "[EXAMPLE".
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = _str_or_none(row[0] if len(row) > 0 else None)
        if key is None:
            continue
        if key.lower() == "key":
            continue
        if key.startswith("pssgen Register"):
            continue
        if _is_example_row(key):
            continue
        value = _str_or_none(row[1] if len(row) > 1 else None)
        result[key] = value if value is not None else ""
    return result


def _read_blocks_sheet(wb) -> list:
    """Extract the Blocks sheet into a list of block dicts."""
    if "Blocks" not in wb.sheetnames:
        return []

    ws = wb["Blocks"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        block_name = _str_or_none(row[0] if len(row) > 0 else None)
        if block_name is None:
            continue
        if block_name.lower() == "block_name":
            continue  # stray header row
        if _is_example_row(block_name):
            continue
        result.append({
            "block_name":      block_name,
            "base_address":    _str_or_none(row[1] if len(row) > 1 else None) or "",
            "data_width_bits": _str_or_none(row[2] if len(row) > 2 else None) or "32",
            "reset_domain":    _str_or_none(row[3] if len(row) > 3 else None) or "",
            "clock_domain":    _str_or_none(row[4] if len(row) > 4 else None) or "",
            "description":     _str_or_none(row[5] if len(row) > 5 else None) or "",
        })
    return result


def _read_registermap_sheet(wb) -> list:
    """Group RegisterMap rows by (block, reg, offset) and return register list."""
    if "RegisterMap" not in wb.sheetnames:
        return []

    ws = wb["RegisterMap"]
    # Collect all valid field rows first
    raw_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        block_name = _str_or_none(row[0])
        field_name = _str_or_none(row[5])
        if block_name is None or field_name is None:
            continue
        if block_name.lower() == "block_name":
            continue
        if _is_example_row(block_name):
            continue
        raw_rows.append(row)

    # Group into registers preserving insertion order
    reg_map: dict[tuple, dict] = {}
    reg_order: list[tuple] = []

    for row in raw_rows:
        block   = _str_or_none(row[0]) or ""
        reg_name = _str_or_none(row[1]) or ""
        reg_desc = _str_or_none(row[2]) or ""
        offset  = _str_or_none(row[3]) or ""
        width   = row[4]

        key = (block, reg_name, offset)
        if key not in reg_map:
            try:
                reg_width = int(width) if width is not None else 32
            except (ValueError, TypeError):
                reg_width = 32
            reg_map[key] = {
                "block":       block,
                "name":        reg_name,
                "description": reg_desc,
                "offset":      offset,
                "width":       reg_width,
                "fields":      [],
            }
            reg_order.append(key)

        field = _parse_field_row(row)
        reg_map[key]["fields"].append(field)

    return [reg_map[k] for k in reg_order]


def _parse_field_row(row) -> dict:
    """Extract and normalise a single field row from the RegisterMap sheet."""
    def _int_val(v, default=0):
        try:
            return int(v) if v is not None else default
        except (ValueError, TypeError):
            return default

    def _bool_yes(v):
        return str(v).strip().upper() == "YES" if v is not None else False

    def _col(idx):
        return row[idx] if idx < len(row) else None

    return {
        "field_name":        _str_or_none(_col(5)) or "",
        "bit_offset":        _int_val(_col(6)),
        "bit_width":         _int_val(_col(7), default=1),
        "access":            _str_or_none(_col(8)) or "NA",
        "reset_value":       _str_or_none(_col(9)) or "0x0",
        "volatile":          _bool_yes(_col(10)),
        "hw_access":         _str_or_none(_col(11)) or "NA",
        "sw_access":         _str_or_none(_col(12)) or "NA",
        "field_kind":        _str_or_none(_col(13)) or "normal",
        "enum_ref":          _str_or_none(_col(14)),
        "uvm_has_coverage":  _bool_yes(_col(15)),
        "req_id":            _str_or_none(_col(16)),
        "pss_action":        _str_or_none(_col(17)),
        "hdl_path":          _str_or_none(_col(18)),
        "description":       _str_or_none(_col(19)) or "",
    }


def _read_enums_sheet(wb) -> dict:
    """Group Enums sheet rows by enum_name and return a name→values dict."""
    if "Enums" not in wb.sheetnames:
        return {}

    ws = wb["Enums"]
    result: dict = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        enum_name = _str_or_none(row[0] if len(row) > 0 else None)
        if enum_name is None:
            continue
        if enum_name.lower() == "enum_name":
            continue
        if _is_example_row(enum_name):
            continue
        try:
            value = int(row[1]) if row[1] is not None else 0
        except (ValueError, TypeError):
            value = 0
        symbol      = _str_or_none(row[2] if len(row) > 2 else None) or ""
        description = _str_or_none(row[3] if len(row) > 3 else None) or ""
        if enum_name not in result:
            result[enum_name] = []
        result[enum_name].append({
            "value":       value,
            "symbol":      symbol,
            "description": description,
        })

    return result


# ---------------------------------------------------------------------------
# Simple-block parser (single-sheet, 15–18 columns)
# ---------------------------------------------------------------------------

# Reserved/unused field name pattern for field_kind detection
_RESERVED_PATTERN = re.compile(r"^(reserved|rsvd|unused)$", re.IGNORECASE)

# Status hw_access values for field_kind detection
_STATUS_HW_ACCESS = {
    "set-on-1", "set-on-read", "set-by-input", "status/clear",
    "rc", "rs", "pulse", "tog", "write-1-to-clear",
}

# Access values that do not warrant coverage
_NO_COVER_ACCESS = {"wo", "na", "w", "na"}
# HW access values that do not warrant coverage
_NO_COVER_HW = {"na", "n/a"}


def _normalise_access(raw: str) -> str:
    """Normalise spreadsheet access strings to canonical forms.

    Mapping:
      R/W, R/W → RW
      R/O, RO  → RO
      W/O, WO, W → WO
      Anything else → as-is stripped uppercase

    Args:
        raw: Raw access string from spreadsheet cell.

    Returns:
        Normalised access string.
    """
    s = raw.strip().upper()
    if s in ("R/W", "RW"):
        return "RW"
    if s in ("R/O", "RO"):
        return "RO"
    if s in ("W/O", "WO", "W"):
        return "WO"
    return s


def _normalise_reset(raw) -> str:
    """Normalise reset value to hex string.

    Rules:
      - int 0 → "0x0"
      - int N → "0xN" (hex string)
      - str already containing "0x" → preserve
      - other str → return as-is stripped

    Args:
        raw: Cell value (int or str).

    Returns:
        Hex string reset value.
    """
    if raw is None:
        return "0x0"
    if isinstance(raw, int):
        if raw == 0:
            return "0x0"
        return hex(raw)
    s = str(raw).strip()
    if not s:
        return "0x0"
    return s


def _field_kind_from_simple(field_name: str, hw_access: str) -> str:
    """Determine field_kind from field name and hw_access.

    Priority:
      1. "reserved" if name matches reserved|rsvd|unused (case-insensitive)
      2. "status" if hw_access is in _STATUS_HW_ACCESS set (case-insensitive)
      3. "normal" otherwise

    Args:
        field_name: Field name string.
        hw_access: Hardware access string.

    Returns:
        One of "reserved", "status", or "normal".
    """
    if _RESERVED_PATTERN.match(field_name.strip()):
        return "reserved"
    if hw_access.strip().lower() in _STATUS_HW_ACCESS:
        return "status"
    return "normal"


def _uvm_coverage_from_simple(access: str, hw_access: str) -> bool:
    """Derive uvm_has_coverage default for a simple_block field.

    Coverage is generated unless:
      - hw_access is NA/N/A (hardware never touches the field)
      - access is WO/NA/W (not readable, so coverage is useless)

    Args:
        access: Normalised access string.
        hw_access: Hardware access string (raw or normalised).

    Returns:
        True if coverage should be generated, False otherwise.
    """
    hw_lower = hw_access.strip().lower()
    acc_lower = access.strip().lower()
    if hw_lower in _NO_COVER_HW:
        return False
    if acc_lower in _NO_COVER_ACCESS:
        return False
    return True


def _parse_inline_enum(
    cell_val: str,
    block_name: str,
    field_name: str,
    enums_dict: dict,
) -> str | None:
    """Parse inline enum string and register it in enums_dict.

    Accepts ``"0=DISABLE,1=ENABLE"`` format. Generates a canonical enum type
    name ``<BLOCK>_<FIELD>_t`` in uppercase. Adds an entry to enums_dict and
    returns the type name, or None if cell_val is blank.

    Args:
        cell_val: Raw cell value from Field Enumerations column.
        block_name: Block name (used in type name prefix).
        field_name: Field name (used in type name suffix).
        enums_dict: Mutable dict to register the enum into.

    Returns:
        Enum type name string, or None if no enum.
    """
    if not cell_val or not str(cell_val).strip():
        return None
    raw = str(cell_val).strip()
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    entries = []
    for part in parts:
        if "=" in part:
            val_str, sym = part.split("=", 1)
            try:
                val = int(val_str.strip())
            except ValueError:
                val = 0
            entries.append({
                "value": val,
                "symbol": sym.strip(),
                "description": "",
            })
    if not entries:
        return None
    enum_name = f"{block_name.upper()}_{field_name.upper()}_t"
    if enum_name not in enums_dict:
        enums_dict[enum_name] = entries
    return enum_name


def _parse_simple_block(xlsx_path: str) -> dict:
    """Parse a single-sheet simple_block register map spreadsheet.

    Accepts 15 or 18 columns. Optional columns 16–18 (base_address, req_id,
    pss_action) are absent in 15-column files; defaults apply. Base address
    is inherited row-to-row from the most recent non-blank value in column 16.
    Default base address is ``"0x0000_0000"`` when the column is absent.

    Column mapping (0-indexed):
      0  Block Name         → block identifier
      1  Register Name      → reg name
      2  Register Offset    → offset (hex str)
      3  Register Width     → width (int)
      4  Register Desc      → description
      5  Field Name         → field_name
      6  Bit Offset         → bit_offset (int)
      7  Bit Width          → bit_width (int)
      8  Access             → access (normalised)
      9  Reset Value        → reset_value (hex str)
      10 Field Description  → description
      11 Volatile           → bool
      12 Hardware Access    → hw_access
      13 Software Access    → sw_access
      14 Field Enumerations → inline enum (parsed)
      15 base_address (opt) → inherited
      16 req_id (opt)       → None if blank
      17 pss_action (opt)   → None if blank

    Args:
        xlsx_path: Path to single-sheet ``.xlsx`` file.

    Returns:
        Normalised register_map dict with keys: globals, blocks,
        registers, enums.

    Raises:
        ParseError: If the file cannot be read or has no data rows.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ParseError(
            "openpyxl is required. Install with: pip install openpyxl"
        ) from exc

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"Cannot open workbook: {exc}") from exc

    # Use first sheet regardless of name
    ws = wb[wb.sheetnames[0]]

    enums_dict: dict = {}
    reg_map: dict[tuple, dict] = {}
    reg_order: list[tuple] = []
    block_base: dict[str, str] = {}   # block_name → base_address
    block_order: list[str] = []

    current_base_address = "0x0000_0000"
    first_data_width: int = 32
    first_block_name: str | None = None

    row_iter = ws.iter_rows(min_row=2, values_only=True)

    for raw_row in row_iter:
        # Require at least 15 columns of data
        if len(raw_row) < 6:
            continue

        block_name  = _str_or_none(raw_row[0])
        reg_name    = _str_or_none(raw_row[1])
        field_name  = _str_or_none(raw_row[5])

        if block_name is None or reg_name is None or field_name is None:
            continue
        if block_name.lower() in ("block name", "block_name"):
            continue  # stray header row

        # Optional columns 16–18 (0-indexed 15–17)
        if len(raw_row) > 15 and raw_row[15] is not None:
            base_val = _str_or_none(raw_row[15])
            if base_val:
                current_base_address = base_val

        req_id     = _str_or_none(raw_row[16]) if len(raw_row) > 16 else None
        pss_action = _str_or_none(raw_row[17]) if len(raw_row) > 17 else None

        # Track blocks in order of first appearance
        if block_name not in block_base:
            block_base[block_name] = current_base_address
            block_order.append(block_name)
        if first_block_name is None:
            first_block_name = block_name

        # Register grouping
        reg_offset  = _str_or_none(raw_row[2]) or ""
        reg_width_raw = raw_row[3]
        try:
            reg_width = int(reg_width_raw) if reg_width_raw is not None else 32
        except (ValueError, TypeError):
            reg_width = 32

        if first_data_width == 32 and reg_width:
            first_data_width = reg_width

        reg_desc = _str_or_none(raw_row[4]) or ""

        key = (block_name, reg_name, reg_offset)
        if key not in reg_map:
            reg_map[key] = {
                "block":       block_name,
                "name":        reg_name,
                "description": reg_desc,
                "offset":      reg_offset,
                "width":       reg_width,
                "fields":      [],
            }
            reg_order.append(key)

        # Field extraction
        def _iv(v, d=0):
            try:
                return int(v) if v is not None else d
            except (ValueError, TypeError):
                return d

        def _vol(v):
            if v is None:
                return False
            sv = str(v).strip().upper()
            return sv in ("YES", "Y", "1", "TRUE")

        raw_access = _str_or_none(raw_row[8]) or "NA"
        access = _normalise_access(raw_access)
        hw_access = _str_or_none(raw_row[12]) or "NA"
        sw_access = _str_or_none(raw_row[13]) or "NA"

        enum_cell = _str_or_none(raw_row[14]) if len(raw_row) > 14 else None
        enum_ref = _parse_inline_enum(enum_cell, block_name, field_name,
                                      enums_dict)

        field_kind = _field_kind_from_simple(field_name, hw_access)
        uvm_cov = _uvm_coverage_from_simple(access, hw_access)

        field = {
            "field_name":       field_name,
            "bit_offset":       _iv(raw_row[6]),
            "bit_width":        _iv(raw_row[7], 1),
            "access":           access,
            "reset_value":      _normalise_reset(raw_row[9]),
            "volatile":         _vol(raw_row[11]),
            "hw_access":        hw_access,
            "sw_access":        sw_access,
            "field_kind":       field_kind,
            "enum_ref":         enum_ref,
            "uvm_has_coverage": uvm_cov,
            "req_id":           req_id,
            "pss_action":       pss_action,
            "hdl_path":         None,
            "description":      _str_or_none(raw_row[10]) or "",
        }
        reg_map[key]["fields"].append(field)

    wb.close()

    # Build blocks list
    blocks_list = []
    for bn in block_order:
        blocks_list.append({
            "block_name":      bn,
            "base_address":    block_base[bn],
            "data_width_bits": str(first_data_width),
            "reset_domain":    "",
            "clock_domain":    "",
            "description":     "",
        })

    globals_dict = {
        "project_name":    first_block_name or "",
        "base_address":    block_base.get(first_block_name, "0x0000_0000")
                           if first_block_name else "0x0000_0000",
        "data_width_bits": str(first_data_width),
        "endianness":      "Little",
        "bus_type":        "",
    }

    return {
        "globals":   globals_dict,
        "blocks":    blocks_list,
        "registers": [reg_map[k] for k in reg_order],
        "enums":     enums_dict,
    }


# ---------------------------------------------------------------------------
# System parser (multi-file, two-sheet spreadsheet)
# ---------------------------------------------------------------------------

def _parse_system(xlsx_path: str) -> dict:
    """Parse a system spreadsheet that references multiple block files.

    Reads the ``System`` sheet for project globals and the ``Blocks`` sheet
    for block file references. Resolves each ``spreadsheet_file`` path
    relative to the directory containing ``xlsx_path``, then calls
    ``parse_regmap()`` recursively on each block file.

    Base address values from the Blocks sheet override any base address
    embedded in the referenced block file.

    Merge strategy:
      globals:   from System sheet
      blocks:    union of all blocks lists across all referenced files
      registers: concatenation preserving order
      enums:     union of all enums dicts

    Args:
        xlsx_path: Path to the system-format ``.xlsx`` file.

    Returns:
        Combined register_map dict.

    Raises:
        FileNotFoundError: If a referenced block file cannot be found.
        ParseError: If a block file cannot be parsed.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ParseError(
            "openpyxl is required. Install with: pip install openpyxl"
        ) from exc

    base_dir = os.path.dirname(os.path.abspath(xlsx_path))

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"Cannot open system workbook: {exc}") from exc

    # --- Read System sheet (project globals) ---
    globals_dict: dict = {}
    if "System" in wb.sheetnames:
        ws_sys = wb["System"]
        for row in ws_sys.iter_rows(min_row=2, values_only=True):
            key = _str_or_none(row[0] if len(row) > 0 else None)
            if key is None or key.lower() == "key":
                continue
            val = _str_or_none(row[1] if len(row) > 1 else None)
            globals_dict[key] = val if val is not None else ""

    # --- Read Blocks sheet ---
    if "Blocks" not in wb.sheetnames:
        wb.close()
        raise ParseError(f"System workbook '{xlsx_path}' has no Blocks sheet")

    ws_blk = wb["Blocks"]
    # Parse header row to find column indices
    header_row = next(ws_blk.iter_rows(min_row=1, max_row=1, values_only=True),
                      None)
    if header_row is None:
        wb.close()
        raise ParseError("Blocks sheet has no header row")

    header_lower = [str(c).lower().strip() if c else "" for c in header_row]
    col_idx = {name: i for i, name in enumerate(header_lower)}

    block_entries = []
    for row in ws_blk.iter_rows(min_row=2, values_only=True):
        bn = _str_or_none(row[col_idx.get("block_name", 0)] if col_idx.get("block_name", 0) < len(row) else None)
        if bn is None or bn.lower() in ("block_name", "block name"):
            continue
        sf_idx = col_idx.get("spreadsheet_file", 1)
        sf = _str_or_none(row[sf_idx] if sf_idx < len(row) else None)
        ba_idx = col_idx.get("base_address", 2)
        ba = _str_or_none(row[ba_idx] if ba_idx < len(row) else None)
        desc_idx = col_idx.get("description", 3)
        desc = _str_or_none(row[desc_idx] if desc_idx < len(row) else None)
        block_entries.append({
            "block_name":       bn,
            "spreadsheet_file": sf,
            "base_address":     ba,
            "description":      desc or "",
        })

    wb.close()

    # --- Load and merge each referenced block file ---
    merged_blocks: list = []
    merged_registers: list = []
    merged_enums: dict = {}

    for entry in block_entries:
        sf = entry["spreadsheet_file"]
        if sf is None:
            continue
        block_file = os.path.join(base_dir, sf)
        if not os.path.isfile(block_file):
            raise FileNotFoundError(
                f"Block file '{sf}' not found relative to '{base_dir}'"
            )
        block_data = parse_regmap(block_file)

        # Apply system-level base_address override to each block
        if entry["base_address"]:
            for blk in block_data.get("blocks", []):
                blk["base_address"] = entry["base_address"]

        merged_blocks.extend(block_data.get("blocks", []))
        merged_registers.extend(block_data.get("registers", []))
        # Merge enums (first definition wins for name conflicts)
        for ename, evals in block_data.get("enums", {}).items():
            if ename not in merged_enums:
                merged_enums[ename] = evals

    return {
        "globals":   globals_dict,
        "blocks":    merged_blocks,
        "registers": merged_registers,
        "enums":     merged_enums,
    }


# ---------------------------------------------------------------------------
# Plain English register map parser
# ---------------------------------------------------------------------------

# Patterns for plain English format
_REG_HEADING = re.compile(
    r"^\s+(?P<name>[A-Za-z_][A-Za-z0-9_]*)\s+register\s+at\s+offset\s+"
    r"(?P<offset>0x[0-9A-Fa-f_]+)"
    r"(?P<flags>[^:]*):",
    re.IGNORECASE,
)
_FIELD_LINE = re.compile(
    r"^\s+(?P<fname>[A-Za-z_][A-Za-z0-9_]*)\s+field\s+\[(?P<msb>\d+):(?P<lsb>\d+)\]\s+"
    r"(?P<access>[A-Z0-9]+)"
    r"(?:\s+reset=(?P<reset>\S+))?"
    r"(?:\s+[—\-]+\s+(?P<desc>.*))?",
    re.IGNORECASE,
)
_SECTION_HEADING = re.compile(r"^[A-Za-z].*:\s*$")


def _parse_intent_regmap(content: str) -> dict:
    """Parse a ``register map:`` section from plain English intent content.

    Handles ``.intent`` file content that contains a ``register map:``
    section with register and field descriptions. This is Tier 2 — metadata
    is limited and no enum definitions are produced.

    The function finds the ``register map:`` section, stops at the next
    top-level section heading (a line ending in ``:`` with no leading
    whitespace), and parses registers and fields from the section body.

    Args:
        content: Full text of the intent file or a partial string containing
                 a ``register map:`` section.

    Returns:
        Normalised register_map dict with minimal globals, one DEFAULT block,
        parsed registers, and empty enums.
    """
    lines = content.splitlines()

    # Find the register map: section
    in_regmap = False
    regmap_lines: list[str] = []

    for line in lines:
        stripped = line.rstrip()
        # Detect section start (top-level, no leading whitespace, ends with :)
        if _SECTION_HEADING.match(stripped):
            heading_lower = stripped.lower().rstrip(":")
            if heading_lower.strip() == "register map":
                in_regmap = True
                continue
            elif in_regmap:
                # A new top-level section — stop collecting
                break
        if in_regmap:
            regmap_lines.append(stripped)

    registers: list[dict] = []
    current_reg: dict | None = None

    for line in regmap_lines:
        if not line.strip():
            continue

        reg_match = _REG_HEADING.match(line)
        if reg_match:
            current_reg = {
                "block":       "DEFAULT",
                "name":        reg_match.group("name").upper(),
                "description": f"{reg_match.group('name')} register",
                "offset":      reg_match.group("offset"),
                "width":       32,
                "fields":      [],
            }
            # Mark all fields volatile if "(volatile)" in flags
            flags = reg_match.group("flags") or ""
            current_reg["_volatile_default"] = "volatile" in flags.lower()
            registers.append(current_reg)
            continue

        field_match = _FIELD_LINE.match(line)
        if field_match and current_reg is not None:
            msb = int(field_match.group("msb"))
            lsb = int(field_match.group("lsb"))
            bit_offset = lsb
            bit_width  = msb - lsb + 1
            access = field_match.group("access").upper()
            reset_val = field_match.group("reset") or "0x0"
            description = (field_match.group("desc") or "").strip()
            is_volatile = current_reg.get("_volatile_default", False)

            current_reg["fields"].append({
                "field_name":       field_match.group("fname").upper(),
                "bit_offset":       bit_offset,
                "bit_width":        bit_width,
                "access":           access,
                "reset_value":      reset_val,
                "volatile":         is_volatile,
                "hw_access":        "NA",
                "sw_access":        access,
                "field_kind":       "normal",
                "enum_ref":         None,
                "uvm_has_coverage": True,
                "req_id":           None,
                "pss_action":       None,
                "hdl_path":         None,
                "description":      description,
            })

    # Strip internal parsing scratch key
    for reg in registers:
        reg.pop("_volatile_default", None)

    # Derive block list from parsed registers
    blocks = [{"block_name": "DEFAULT", "base_address": "0x0",
               "data_width_bits": "32", "reset_domain": "",
               "clock_domain": "", "description": ""}] if registers else []

    return {
        "globals": {},
        "blocks":  blocks,
        "registers": registers,
        "enums": {},
    }
