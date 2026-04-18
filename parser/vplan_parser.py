# ===========================================================
# FILE:         parser/vplan_parser.py
# PROJECT:      pssgen — AI-Driven PSS + UVM + C Testbench Generator
# COPYRIGHT:    Copyright (c) 2026 BelTech Systems LLC
# LICENSE:      MIT License — see LICENSE file for details
# ===========================================================
#
# DESCRIPTION:
#   Reads the VPR (Verification Planning and Results) spreadsheet and returns
#   a VplanParseResult that duck-types both ReqParseResult and IntentParseResult.
#   Replaces req_parser.py and intent_parser.py as the source of requirements and
#   coverage intent when a .xlsx VPR is provided (D-031). Reads two tabs: VPR
#   (requirements + dispositions) and Coverage_Goals (COV item definitions).
#
# LAYER:        1 — parser
# PHASE:        v6a
#
# FUNCTIONS:
#   parse_vplan(vplan_file)
#     Parse a VPR spreadsheet and return a VplanParseResult.
#
# DEPENDENCIES:
#   Standard library:  dataclasses, typing
#   Internal:          none
#   Third-party:       openpyxl
#
# HISTORY:
#   v6a  2026-04-12  SB  Initial implementation; replaces req_parser + intent_parser (OI-30, D-031)
#
# ===========================================================
"""parser/vplan_parser.py — VPR spreadsheet parser.

Phase: v6a
Layer: 1 (parser)

Reads the VPR spreadsheet (.xlsx) and produces a VplanParseResult that
duck-types both ReqParseResult and IntentParseResult so that gap_agent,
pss_gen, and all downstream code require no changes below the import swap
in orchestrator.py.
"""
from dataclasses import dataclass, field
from typing import Optional

import openpyxl


# ── VPR tab column indices (0-based, from header row 2) ──────────────────────
_VPR_REQ_ID          = 0
_VPR_FAMILY          = 1
_VPR_STATEMENT       = 3
_VPR_VERIF_METHOD    = 6
_VPR_COVERED_BY      = 7
_VPR_DISPOSITION     = 10
_VPR_WAIVER_RATIONALE = 11
_VPR_RTL_STATUS      = 18

# VPR data starts at row 4 (1-based): row1=group headers, row2=col headers, row3=example
_VPR_DATA_START_ROW = 4

# Req_ID values that indicate a non-data row in the VPR tab
_VPR_SKIP_IDS = {"Req_ID", "[BLOCK-FAM-NNN]"}

# ── Coverage_Goals tab column indices (0-based, from header row 1) ────────────
_COV_ID               = 0
_COV_NAME             = 1
_COV_DESCRIPTION      = 2
_COV_STIMULUS         = 3
_COV_BOUNDARY         = 4
_COV_LINKED_REQS      = 5
_COV_STATUS           = 6
_COV_COVERAGE_TYPE    = 7
_COV_NOTES            = 8
_COV_SEQ_STATUS       = 9
_COV_STIMULUS_VSL     = 10

# Coverage_Goals data starts at row 2 (1-based): row1=headers
_COV_DATA_START_ROW = 2


def _cell_str(row: tuple, col: int) -> str:
    """Return the string value of a cell, or '' if absent or None."""
    if col >= len(row):
        return ""
    val = row[col].value
    if val is None:
        return ""
    return str(val).strip()


def _split_csv(text: str) -> list[str]:
    """Split a comma-separated string into a stripped, non-empty list."""
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def _derive_scheme(req_id: str) -> Optional[str]:
    """Return the scheme prefix of a requirement ID (all but the last segment).

    Args:
        req_id: A requirement ID such as "UART-BR-001".

    Returns:
        Scheme prefix e.g. "UART-BR", or None if fewer than two segments.
    """
    parts = req_id.split("-")
    if len(parts) < 2:
        return None
    return "-".join(parts[:-1])


@dataclass
class VplanParseResult:
    """Result of parsing a VPR spreadsheet.

    Duck-types both ReqParseResult and IntentParseResult so downstream code
    (gap_agent, pss_gen, orchestrator) requires no changes.

    Attributes:
        requirements: Mapping from Req_ID to requirement detail dict.
            Keys: statement (str), verification (list[str]), waived (bool),
            waiver_reason (str), family (str), covered_by (str), rtl_status (str).
        waivers: List of Req_IDs where Disposition == "WAIVED".
        cov_items: Mapping from COV ID to coverage item detail dict.
            Keys: name, description, stimulus_strategy, boundary_values,
            linked_requirements (list[str]), status, coverage_type, notes,
            seq_status (str, default "NONE"), stimulus_vsl (str, default "").
        req_ids: Non-waived Req_IDs in row order.
        req_schemes: Unique scheme prefixes derived from req_ids.
        sections: Always empty dict — IntentParseResult duck-type.
        inline_requirements: Always empty dict — IntentParseResult duck-type.
        intent_waivers: One entry per waived req for IntentParseResult duck-type.
            Each entry: {item: req_id, reason: waiver_reason, req_ids: [req_id]}.
    """
    requirements: dict[str, dict] = field(default_factory=dict)
    waivers: list[str] = field(default_factory=list)
    cov_items: dict[str, dict] = field(default_factory=dict)
    req_ids: list[str] = field(default_factory=list)
    req_schemes: list[str] = field(default_factory=list)
    sections: dict = field(default_factory=dict)
    inline_requirements: dict = field(default_factory=dict)
    intent_waivers: list[dict] = field(default_factory=list)

    @property
    def mode(self) -> str:
        """Classify as "full" or "campaign".

        Returns:
            "full" if at least one non-waived requirement exists;
            "campaign" if all entries are waived or no entries are present.
        """
        if not self.requirements:
            return "campaign"
        if all(entry.get("waived") for entry in self.requirements.values()):
            return "campaign"
        return "full"


def parse_vplan(vplan_file: str) -> VplanParseResult:
    """Parse a VPR spreadsheet and return a VplanParseResult.

    Reads the VPR tab (requirements, dispositions, coverage cross-references)
    and the Coverage_Goals tab (COV item definitions). The result duck-types
    both ReqParseResult and IntentParseResult for transparent downstream use.

    Args:
        vplan_file: Path to the VPR spreadsheet (.xlsx).

    Returns:
        VplanParseResult containing requirements, coverage items, and all
        fields required for duck-typing downstream parsers.

    Raises:
        ValueError: If the VPR or Coverage_Goals tab is missing.
        FileNotFoundError: If vplan_file does not exist.
    """
    wb = openpyxl.load_workbook(vplan_file, read_only=True, data_only=True)

    if "VPR" not in wb.sheetnames:
        raise ValueError(f"VPR tab not found in {vplan_file}")
    if "Coverage_Goals" not in wb.sheetnames:
        raise ValueError(f"Coverage_Goals tab not found in {vplan_file}")

    requirements: dict[str, dict] = {}
    waivers: list[str] = []
    req_ids_ordered: list[str] = []

    ws_vpr = wb["VPR"]
    for row_idx, row in enumerate(ws_vpr.iter_rows(), start=1):
        if row_idx < _VPR_DATA_START_ROW:
            continue

        req_id = _cell_str(row, _VPR_REQ_ID)
        if not req_id or req_id in _VPR_SKIP_IDS:
            continue

        family         = _cell_str(row, _VPR_FAMILY)
        statement      = _cell_str(row, _VPR_STATEMENT)
        verif_raw      = _cell_str(row, _VPR_VERIF_METHOD)
        covered_by     = _cell_str(row, _VPR_COVERED_BY)
        disposition    = _cell_str(row, _VPR_DISPOSITION)
        waiver_reason  = _cell_str(row, _VPR_WAIVER_RATIONALE)
        rtl_status     = _cell_str(row, _VPR_RTL_STATUS)

        is_waived = disposition.upper() == "WAIVED"
        verification = _split_csv(verif_raw)

        requirements[req_id] = {
            "statement":      statement,
            "verification":   verification,
            "waived":         is_waived,
            "waiver_reason":  waiver_reason,
            "family":         family,
            "covered_by":     covered_by,
            "rtl_status":     rtl_status,
        }

        if is_waived:
            waivers.append(req_id)
        else:
            req_ids_ordered.append(req_id)

    # ── Coverage_Goals tab ────────────────────────────────────────────────────
    cov_items: dict[str, dict] = {}

    ws_cov = wb["Coverage_Goals"]
    for row_idx, row in enumerate(ws_cov.iter_rows(), start=1):
        if row_idx < _COV_DATA_START_ROW:
            continue

        cov_id = _cell_str(row, _COV_ID)
        if not cov_id.startswith("COV-"):
            continue

        linked_raw = _cell_str(row, _COV_LINKED_REQS)

        cov_items[cov_id] = {
            "name":               _cell_str(row, _COV_NAME),
            "description":        _cell_str(row, _COV_DESCRIPTION),
            "stimulus_strategy":  _cell_str(row, _COV_STIMULUS),
            "boundary_values":    _cell_str(row, _COV_BOUNDARY),
            "linked_requirements": _split_csv(linked_raw),
            "status":             _cell_str(row, _COV_STATUS),
            "coverage_type":      _cell_str(row, _COV_COVERAGE_TYPE),
            "notes":              _cell_str(row, _COV_NOTES),
            "seq_status":         _cell_str(row, _COV_SEQ_STATUS) or "NONE",
            "stimulus_vsl":       _cell_str(row, _COV_STIMULUS_VSL),
        }

    wb.close()

    # ── Derive schemes ────────────────────────────────────────────────────────
    scheme_seen: set[str] = set()
    req_schemes: list[str] = []
    for rid in req_ids_ordered:
        scheme = _derive_scheme(rid)
        if scheme and scheme not in scheme_seen:
            scheme_seen.add(scheme)
            req_schemes.append(scheme)

    # ── intent_waivers — duck-type IntentParseResult.waivers ─────────────────
    intent_waivers: list[dict] = []
    for wid in waivers:
        entry = requirements.get(wid, {})
        intent_waivers.append({
            "item":    wid,
            "reason":  entry.get("waiver_reason", ""),
            "req_ids": [wid],
        })

    return VplanParseResult(
        requirements=requirements,
        waivers=waivers,
        cov_items=cov_items,
        req_ids=req_ids_ordered,
        req_schemes=req_schemes,
        sections={},
        inline_requirements={},
        intent_waivers=intent_waivers,
    )
