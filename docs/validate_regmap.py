"""
Validates counter_regmap.xlsx against the specification.
Run from project root with venv active:
    python docs/validate_regmap.py
"""
import re
import sys
from openpyxl import load_workbook

FIXTURE = "tests/fixtures/counter_regmap.xlsx"

def check(cond, msg):
    if not cond:
        print(f"  FAIL: {msg}")
        return False
    print(f"  OK:   {msg}")
    return True

def main():
    wb = load_workbook(FIXTURE, data_only=True)
    failures = 0

    # ── Globals ───────────────────────────────────────────────────────────────
    print("\n=== Globals ===")
    ws = wb["Globals"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                 if any(c is not None for c in r)]
    failures += 0 if check(len(data_rows) == 9,
                            f"9 data rows (got {len(data_rows)})") else 1

    # ── Blocks ────────────────────────────────────────────────────────────────
    print("\n=== Blocks ===")
    ws = wb["Blocks"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                 if any(c is not None for c in r)]
    failures += 0 if check(len(data_rows) == 1,
                            f"1 data row (got {len(data_rows)})") else 1
    if data_rows:
        failures += 0 if check(data_rows[0][0] == "COUNTER",
                                f"block_name = COUNTER") else 1

    # ── RegisterMap ───────────────────────────────────────────────────────────
    print("\n=== RegisterMap ===")
    ws = wb["RegisterMap"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                 if any(c is not None for c in r)]
    failures += 0 if check(len(data_rows) == 15,
                            f"15 field rows (got {len(data_rows)})") else 1

    req_re = re.compile(r"^FUNC-REQ-\d+$")
    bad_req = []
    blank_field = []
    bad_cov = []

    for i, row in enumerate(data_rows, start=1):
        field_name       = row[5]   # col 6
        uvm_has_coverage = row[15]  # col 16
        req_id           = row[16]  # col 17

        if not field_name:
            blank_field.append(i)
        if req_id and not req_re.match(str(req_id)):
            bad_req.append((i, req_id))
        if uvm_has_coverage not in ("YES", "NO"):
            bad_cov.append((i, uvm_has_coverage))

    failures += 0 if check(len(blank_field) == 0,
                            f"No blank field_name values (blank rows: {blank_field})") else 1
    failures += 0 if check(len(bad_req) == 0,
                            f"All req_ids match FUNC-REQ-NNN (bad: {bad_req})") else 1
    failures += 0 if check(len(bad_cov) == 0,
                            f"All uvm_has_coverage are YES/NO (bad: {bad_cov})") else 1

    # Row count by register
    reg_counts = {}
    for row in data_rows:
        reg = row[1]
        reg_counts[reg] = reg_counts.get(reg, 0) + 1
    print(f"  INFO: Field rows by register: {reg_counts}")
    expected = {"CTRL": 3, "STATUS": 4, "COUNT": 2, "LOAD": 2, "INT_CTRL": 4}
    for reg, exp in expected.items():
        got = reg_counts.get(reg, 0)
        failures += 0 if check(got == exp,
                                f"  {reg}: {exp} fields (got {got})") else 1

    # ── Enums ─────────────────────────────────────────────────────────────────
    print("\n=== Enums ===")
    ws = wb["Enums"]
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                 if any(c is not None for c in r)]
    failures += 0 if check(len(data_rows) == 4,
                            f"4 data rows (got {len(data_rows)})") else 1

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'='*40}")
    if failures == 0:
        print("ALL CHECKS PASSED")
    else:
        print(f"{failures} CHECK(S) FAILED")
    return failures

if __name__ == "__main__":
    sys.exit(main())
