"""
gen_vpr_v2.py — VPR template v2 structural update.

New columns: Verification_Category, Applicability (Group B after Risk_Acceptance),
             Config_Set, Tool_Version (Group D after Gate_Evidence).
New CF state: AT_RISK (orange).
New Sim_Mode values: post-layout, HIL.
Coverage_Goals: Coverage_Type column.
Anomalies: Impact, Regression_Fixed columns.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = 'docs/pssgen_vpr_template.xlsx'
BALU_PATH     = 'ip/buffered_axi_lite_uart/buffered_axi_lite_uart_vplan.xlsx'

# ── Column layout AFTER all insertions (1-indexed) ───────────────────────────
# Insert 2 at col 14 (after Risk_Acceptance M=13)
# Insert 2 at col 25 (after Gate_Evidence X=24)
#
# 1-13 : A-M  unchanged
# 14   : N    Verification_Category  (NEW)
# 15   : O    Applicability          (NEW)
# 16-24: P-X  old 14-22  (Test_Name … Gate_Evidence)
# 25   : Y    Config_Set             (NEW)
# 26   : Z    Tool_Version           (NEW)
# 27-34: AA-AH  old 23-30  (Anomaly_ID … Spec_Revision)
#
# Key refs: K=Disposition, H=Covered_By, S(19)=RTL_Status,
#           W(23)=Gate_Status, AA(27)=Anomaly_ID,
#           AB(28)=Overall_Status, AC(29)=Closure_Author

# After inserts, old group start cols shift to:
# GROUP A  1→1   GROUP B  7→7   GROUP C  14→16  GROUP D  17→19
# GROUP E  24→28  GROUP F  28→32

GROUP_HEADERS = [
    # (new_start, new_end, label, fill_hex)
    (1,  6,  'GROUP A \u2014 REQUIREMENT IDENTITY  (pssgen import \u2014 do not edit)', 'D9E1F2'),
    (7,  15, 'GROUP B \u2014 VERIFICATION PLANNING  (engineer)',                        'E2EFDA'),
    (16, 18, 'GROUP C \u2014 IMPLEMENTATION  (engineer)',                               'FFF2CC'),
    (19, 27, 'GROUP D \u2014 EXECUTION RESULTS  (pssgen)',                              'FCE4D6'),
    (28, 31, 'GROUP E \u2014 CLOSURE  (engineer)',                                      'F2F2F2'),
    (32, 34, 'GROUP F \u2014 PROGRAM CONTEXT  (set once per release)',                  'EAD1DC'),
]

def safe_unmerge_row1(ws):
    """Remove all row-1 merged ranges and clear the MergedCell stubs from _cells."""
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row == 1:
            ws.merged_cells.ranges.remove(mc)
            for r, c in mc.cells:
                try:
                    del ws._cells[(r, c)]
                except KeyError:
                    pass


VER_CAT = {
    'PAR':  'Parametric',
    'IF':   'Protocol_Compliance',
    'EN':   'Functional',
    'BR':   'Functional',
    'FF':   'Functional',
    'FIFO': 'Functional',
    'TO':   'Functional',
    'INT':  'Functional',
    'REG':  'Structural',
    'RST':  'Reset',
    'VER':  'Infrastructure',
}

COV_TYPE = {
    'COV-001':'Functional','COV-002':'Functional','COV-003':'Functional',
    'COV-004':'Functional','COV-005':'Functional','COV-006':'Functional',
    'COV-007':'Functional','COV-008':'Functional','COV-009':'Functional',
    'COV-010':'Functional','COV-011':'Functional','COV-012':'Functional',
    'COV-013':'Functional','COV-014':'Functional','COV-015':'Functional',
    'COV-016':'Functional','COV-017':'Structural', 'COV-018':'Structural',
    'COV-019':'Functional',
}

def overall_formula(r):
    return (
        f'=IF(K{r}="WAIVED","WAIVED",'
        f'IF(H{r}="","OPEN",'
        f'IF(S{r}="BLOCKED","BLOCKED",'
        f'IF(S{r}="FAIL","FAILING",'
        f'IF(S{r}="PASS",'
        f'IF(AA{r}<>"","AT_RISK",'
        f'IF(AC{r}<>"","CLOSED","PASSING")),'
        f'"COVERED")))))'
    )

# ── VPR sheet ────────────────────────────────────────────────────────────────
def update_vpr(ws, is_balu=False):
    df = Font(name='Arial', size=9)
    gf = Font(name='Arial', size=9, italic=True, color='595959')
    gfill = PatternFill('solid', fgColor='F9F9F9')
    hf = Font(name='Arial', size=9, bold=True)

    # Step 1: Remove row 1 merged ranges and clear MergedCell stubs
    safe_unmerge_row1(ws)

    # Step 2: Insert columns (cell values shift automatically)
    ws.insert_cols(14, 2)   # Verification_Category, Applicability
    ws.insert_cols(25, 2)   # Config_Set, Tool_Version

    # Step 3: Re-merge row 1 at new ranges (values already shifted to correct cols)
    for start_col, end_col, label, fill_hex in GROUP_HEADERS:
        # Overwrite the label text (some shifted labels retain old text - update them)
        cell = ws.cell(row=1, column=start_col)
        cell.value = label
        cell.font  = hf
        cell.fill  = PatternFill('solid', fgColor=fill_hex)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=end_col)

    # Step 4: Set row 2 headers and column widths for new columns
    new_cols = {
        14: ('Verification_Category', 20),
        15: ('Applicability',         20),
        25: ('Config_Set',            30),
        26: ('Tool_Version',          18),
    }
    for col, (hdr, width) in new_cols.items():
        c = ws.cell(row=2, column=col, value=hdr)
        c.font      = hf
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    # Step 5: Example row 3 — new column values + updated formula
    example_vals = {
        14: 'Functional',
        15: 'ALL',
        25: 'G_FIFO_DEPTH=16, G_DEFAULT_BAUD=115200, G_CLK_FREQ_HZ=100000000',
        26: 'Vivado 2024.1 / Questa 23.2',
    }
    for col, val in example_vals.items():
        c = ws.cell(row=3, column=col, value=val)
        c.font      = gf
        c.fill      = gfill
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    f3 = ws.cell(row=3, column=28, value=overall_formula(3))
    f3.font      = gf
    f3.fill      = gfill
    f3.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    # Step 6: Rebuild data validations
    ws.data_validations.dataValidation = []
    dvs = [
        ('C4:C2000', '"functional,structural,parametric,build-time,infrastructure"'),
        ('G4:G2000', '"test,inspection,analysis,build-time"'),
        ('K4:K2000', '"GENERATED,CONFIRMED,WAIVED"'),
        ('M4:M2000', '"LOW,MEDIUM,HIGH"'),
        ('N4:N2000', '"Functional,Structural,Protocol_Compliance,Reset,CDC,Error_Handling,Parametric,Infrastructure"'),
        ('O4:O2000', '"ALL,PARITY_ENABLED,FIFO_DEPTH_GT_16,CONFIG_SPECIFIC"'),
        ('R4:R2000', '"RTL,gate-level,post-layout,HIL"'),          # was P, now R
        ('S4:S2000', '"PASS,FAIL,NOT_RUN,BLOCKED"'),               # was Q, now S
        ('W4:W2000', '"PASS,FAIL,NOT_RUN"'),                        # was U, now W
    ]
    for sqref, formula1 in dvs:
        ws.add_data_validation(
            DataValidation(type='list', formula1=formula1, sqref=sqref, allow_blank=True)
        )

    # Step 7: Rebuild conditional formatting
    ws.conditional_formatting = ConditionalFormattingList()
    cf_range = 'A4:AH2000'
    cf_rules = [
        ('$AB4="CLOSED"',  'C6EFCE', 1),   # green
        ('$AB4="PASSING"', 'FFEB9C', 2),   # yellow
        ('$AB4="AT_RISK"', 'FFAB40', 3),   # orange (new)
        ('$AB4="FAILING"', 'FFC7CE', 4),   # red
        ('$AB4="WAIVED"',  'D9D9D9', 5),   # grey
    ]
    for formula, fill_hex, priority in cf_rules:
        rule = FormulaRule(
            formula=[formula],
            fill=PatternFill('solid', fgColor=fill_hex),
            stopIfTrue=False
        )
        rule.priority = priority
        ws.conditional_formatting.add(cf_range, rule)

    # Step 8: If BALU, populate new columns + update all formulas
    if is_balu:
        for row in range(4, 145):
            req_id = ws.cell(row=row, column=1).value
            if not req_id:
                continue
            fam = req_id.split('-')[1] if '-' in req_id else ''
            vc = VER_CAT.get(fam, 'Functional')

            c14 = ws.cell(row=row, column=14, value=vc)
            c14.font = df
            c14.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            c15 = ws.cell(row=row, column=15, value='ALL')
            c15.font = df
            c15.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Update Overall_Status formula (now at col 28)
            cf = ws.cell(row=row, column=28, value=overall_formula(row))
            cf.font      = df
            cf.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    print(f"  VPR done. Max col = {ws.max_column}")


# ── Coverage_Goals sheet ─────────────────────────────────────────────────────
def update_coverage_goals(ws, is_balu=False):
    df = Font(name='Arial', size=9)
    hf = Font(name='Arial', size=9, bold=True)

    # Insert Coverage_Type at col 8 (before Notes)
    ws.insert_cols(8, 1)

    # Fix row 1 merge (A1:H1 → A1:I1)
    safe_unmerge_row1(ws)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    # Header
    c = ws.cell(row=2, column=8, value='Coverage_Type')
    c.font      = hf
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(8)].width = 18

    # Dropdown
    ws.add_data_validation(DataValidation(
        type='list', formula1='"Functional,Code,Assertion,Toggle,FSM,Mixed"',
        sqref='H3:H200', allow_blank=True
    ))

    if is_balu:
        for row in range(3, 22):
            cov_id = ws.cell(row=row, column=1).value
            if cov_id and cov_id in COV_TYPE:
                c = ws.cell(row=row, column=8, value=COV_TYPE[cov_id])
                c.font      = df
                c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    print(f"  Coverage_Goals done. Max col = {ws.max_column}")


# ── Anomalies sheet ──────────────────────────────────────────────────────────
def update_anomalies(ws):
    hf = Font(name='Arial', size=9, bold=True)

    # Fix row 1 merge (A1:J1 → A1:L1)
    safe_unmerge_row1(ws)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    for col, hdr, width in [(11, 'Impact', 18), (12, 'Regression_Fixed', 18)]:
        c = ws.cell(row=2, column=col, value=hdr)
        c.font      = hf
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.add_data_validation(DataValidation(
        type='list', formula1='"Functional,Timing,None"',
        sqref='K3:K1000', allow_blank=True
    ))
    ws.add_data_validation(DataValidation(
        type='list', formula1='"YES,NO,PENDING"',
        sqref='L3:L1000', allow_blank=True
    ))

    print(f"  Anomalies done. Max col = {ws.max_column}")


# ── Summary sheet (BALU only) — fix VPR!X refs → VPR!AB ────────────────────
def update_summary(ws):
    """Overall_Status moved from col X(24) to col AB(28) — update refs."""
    updated = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'VPR!X' in cell.value:
                cell.value = cell.value.replace('VPR!X', 'VPR!AB')
                updated += 1
    print(f"  Summary: updated {updated} formula(s) (VPR!X -> VPR!AB)")


# ── Process a workbook ───────────────────────────────────────────────────────
def process_wb(path, is_balu=False):
    print(f"\nProcessing: {path}")
    wb = openpyxl.load_workbook(path)
    update_vpr(wb['VPR'], is_balu=is_balu)
    update_coverage_goals(wb['Coverage_Goals'], is_balu=is_balu)
    update_anomalies(wb['Anomalies'])
    if is_balu:
        update_summary(wb['Summary'])
    wb.save(path)
    print(f"  Saved: {path}")


# ── Verification ─────────────────────────────────────────────────────────────
def verify(path, is_balu=False):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb['VPR']
    print(f"\n--- Verify: {path} ---")

    col14_hdr = ws.cell(row=2, column=14).value
    col28_hdr = ws.cell(row=2, column=28).value
    print(f"  Col 14 header: {col14_hdr!r}  (expected 'Verification_Category')")
    print(f"  Col 28 header: {col28_hdr!r}  (expected 'Overall_Status')")
    print(f"  VPR max_col:   {ws.max_column}  (expected 34)")

    if is_balu:
        data_rows = sum(1 for r in ws.iter_rows(min_row=4, max_row=144, min_col=1, max_col=1)
                        if r[0].value)
        waived    = sum(1 for r in ws.iter_rows(min_row=4, max_row=144, min_col=11, max_col=11)
                        if r[0].value == 'WAIVED')
        open_nw   = [r[0].value
                     for r in ws.iter_rows(min_row=4, max_row=144)
                     if not r[7].value and r[10].value != 'WAIVED']
        formulas  = sum(1 for r in ws.iter_rows(min_row=3, max_row=144, min_col=28, max_col=28)
                        if r[0].value and str(r[0].value).startswith('='))
        sample    = ws.cell(row=4, column=28).value
        formula_ok = sample and 'S4' in sample and 'AA4' in sample and 'AC4' in sample

        print(f"  Data rows (4..144):  {data_rows}  (expected 141)")
        print(f"  WAIVED rows:         {waived}  (expected 2)")
        print(f"  Open non-waived:     {len(open_nw)}  (expected 1)")
        if open_nw:
            print(f"    Open: {open_nw}")
        print(f"  Formulas (3..144):   {formulas}  (expected 142)")
        print(f"  Formula refs S/AA/AC: {formula_ok}  (expected True)")
        print(f"  Sample formula row4: {sample}")

        ws_cg = wb['Coverage_Goals']
        cg8 = ws_cg.cell(row=2, column=8).value
        print(f"  CG col 8 header:     {cg8!r}  (expected 'Coverage_Type')")

        ws_an = wb['Anomalies']
        an11 = ws_an.cell(row=2, column=11).value
        print(f"  Anomalies col 11:    {an11!r}  (expected 'Impact')")

        ok = (data_rows == 141 and waived == 2 and len(open_nw) == 1 and
              open_nw[0] == 'UART-BR-004' and formulas == 142 and
              formula_ok and col14_hdr == 'Verification_Category' and
              col28_hdr == 'Overall_Status' and ws.max_column == 34)
        print(f"\n  {'ALL CHECKS PASS' if ok else 'CHECKS FAILED'}")
        return ok
    else:
        ok = (col14_hdr == 'Verification_Category' and
              col28_hdr == 'Overall_Status' and ws.max_column == 34)
        print(f"\n  {'ALL CHECKS PASS' if ok else 'CHECKS FAILED'}")
        return ok


if __name__ == '__main__':
    process_wb(TEMPLATE_PATH, is_balu=False)
    process_wb(BALU_PATH,     is_balu=True)
    verify(TEMPLATE_PATH, is_balu=False)
    verify(BALU_PATH,     is_balu=True)
